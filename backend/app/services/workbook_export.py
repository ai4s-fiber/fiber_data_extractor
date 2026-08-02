"""Readable materials-science workbook export.

The workbook is organized by scientific entity and fact granularity instead
of mirroring persistence or extraction-pipeline tables.  Core worksheets
contain only literature, samples, composition, process, structure, and
performance.  Evidence and review metadata live in a dedicated quality sheet.
"""

from __future__ import annotations

import hashlib
import re
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.models.candidate_record import CandidateRecord
from app.models.document_parse import DocumentBlock
from app.models.evidence_item import EvidenceItem
from app.models.fact_candidate import FactCandidate
from app.models.paper import Paper
from app.models.sample_catalog import SampleCatalog
from app.services.material_data_model import (
    SHEET_COLUMNS,
    MaterialDataset,
    build_material_dataset,
)


# Deprecated compatibility contract for code that still imports the legacy
# 40-column definition.  New workbooks do not create a Main_Data sheet.
MAIN_DATA_COLUMNS = [
    "record_id",
    "paper_id",
    "paper_title",
    "doi_or_url",
    "year",
    "journal",
    "sample_group_id",
    "sample_id",
    "material_system",
    "fiber_type",
    "variable_name",
    "variable_value",
    "variable_unit",
    "composition_expression",
    "matrix_name",
    "matrix_content",
    "matrix_unit",
    "additive_expression",
    "solvent_or_aid",
    "composition_evidence",
    "process_route",
    "spinning_method",
    "process_parameters",
    "post_treatment",
    "process_evidence",
    "structure_methods",
    "structure_features",
    "structure_evidence",
    "performance_category",
    "performance_metric",
    "performance_value",
    "performance_unit",
    "performance_method",
    "performance_condition",
    "performance_evidence",
    "extraction_method",
    "evidence_text",
    "ai_confidence",
    "review_status",
    "reviewer_comment",
]

WORKBOOK_SHEET_COLUMNS = SHEET_COLUMNS
MAX_DATA_ROWS_PER_SHEET = 1_000_000

MASTER_DATA_SHEET = "01_数据主表"
MASTER_DATA_COLUMNS = [
    "数据ID",
    "文献编号",
    "论文题目",
    "DOI|URL",
    "年份",
    "期刊|会议",
    "材料类别",
    "具体材料对象|样品编号",
    "纤维形态",
    "原料|前驱体|基体",
    "增强|填料|改性组分",
    "成分配比|浓度",
    "溶剂|助剂",
    "工艺路线",
    "关键工艺参数",
    "后处理条件",
    "结构表征方法",
    "结构指标名称",
    "结构数值",
    "结构单位",
    "性能测试方法|标准",
    "性能指标名称",
    "性能数值",
    "性能单位",
    "测试条件",
    "结果描述|结论",
    "数据来源位置",
    "原文图表编号",
    "是否完整",
    "缺失信息说明",
    "备注",
]

_MASTER_GROUPS = [
    ("文献与样品", 1, 9, "1F4E78", "D9EAF7"),
    ("成分", 10, 13, "2E7D32", "E2F0D9"),
    ("工艺", 14, 16, "A05A16", "FCE4D6"),
    ("结构", 17, 20, "5E35B1", "E4DFEC"),
    ("性能", 21, 25, "B3261E", "FDE9E7"),
    ("结论与证据", 26, 31, "455A64", "E7E6E6"),
]

_GENERIC_PROCESS_STAGE_KEYS = frozenset({
    "process",
    "processing",
    "加工",
    "工艺",
    "工艺参数",
    "总体路线",
    "成形",
})

_SHEET_COLORS = {
    "01_文献": "1F4E78",
    "02_样品总览": "0F6B5B",
    "03_成分": "2E7D32",
    "04_工艺": "A05A16",
    "05_结构": "5E35B1",
    "06_性能": "B3261E",
    "90_证据与质控": "455A64",
}

_ROW_MEANINGS = {
    "01_文献": "一行代表一篇文献",
    "02_样品总览": "一行代表文献中的一个实际样品或实验条件",
    "03_成分": "一行代表一个样品—组分关系",
    "04_工艺": "一行代表一个工艺步骤或参数",
    "05_结构": "一行代表一个结构/表征事实",
    "06_性能": "一行代表一个性能测量事实",
    "90_证据与质控": "一行代表一条事实的原文证据或质控记录",
}


def generate_structured_workbook(
    *,
    records: list[CandidateRecord],
    papers: list[Paper],
    evidence_items: list[EvidenceItem],
    document_blocks: list[DocumentBlock],
    filepath: str,
    fact_candidates: list[FactCandidate] | None = None,
    sample_catalogs: list[SampleCatalog] | None = None,
) -> None:
    """Generate the atomic materials workbook and replace the target safely.

    ``document_blocks`` remains in the signature for API/CLI compatibility,
    but parser diagnostics are intentionally excluded from the scientific
    workbook.  Relevant source locations are already carried by facts and
    evidence items.
    """

    dataset = build_material_dataset(
        records=records,
        papers=papers,
        fact_candidates=fact_candidates or [],
        sample_catalogs=sample_catalogs or [],
        evidence_items=evidence_items,
        document_blocks=document_blocks,
    )

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    workbook.properties.title = "AI4S 材料成分—工艺—结构—性能数据"
    workbook.properties.subject = "文献材料事实的原子化结构化导出"
    workbook.properties.creator = "AI4S"
    workbook.properties.keywords = "materials, composition, process, structure, performance"

    master_rows = _build_master_rows(dataset)
    _write_master_sheet(workbook, master_rows)
    _write_readme(workbook, dataset)
    for table_index, (sheet_name, rows) in enumerate(
        dataset.sheet_rows().items(),
        start=1,
    ):
        _write_sheet(
            workbook,
            sheet_name,
            WORKBOOK_SHEET_COLUMNS[sheet_name],
            rows,
            _SHEET_COLORS[sheet_name],
            table_index,
        )

    output_path = Path(filepath)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(
        f".{output_path.stem}-{uuid.uuid4().hex}.tmp{output_path.suffix}"
    )
    try:
        workbook.save(temporary_path)
        temporary_path.replace(output_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _write_readme(workbook: Workbook, dataset: MaterialDataset) -> None:
    ws = workbook.create_sheet("00_说明")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "17365D"
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "AI4S 材料成分—工艺—结构—性能数据工作簿"
    title.font = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="17365D")
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    generated_at = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")
    summary = [
        (
            "主表粒度",
            "01_数据主表沿用原模板逻辑：一行代表一个实际材料样品，"
            "按文献与样品—成分—工艺—结构—性能顺序横向展开",
        ),
        ("明细粒度", "成分、工艺、结构、性能明细表中每个核心事实独占一行"),
        ("核心范围", "材料成分、制备工艺、结构表征、材料性能"),
        ("证据位置", "原文、页码、置信度与复核状态统一放在 90_证据与质控"),
        ("生成时间", generated_at),
        ("去重规则", "同文献、同样品、同指标、同数值、同单位的重复抽取合并"),
        ("缺失值约定", "空单元格表示原文未报告或当前数据中不可确认；不填 0"),
    ]
    for row_index, (label, value) in enumerate(summary, start=3):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=value)
        ws.merge_cells(
            start_row=row_index,
            start_column=2,
            end_row=row_index,
            end_column=4,
        )
        ws.cell(row=row_index, column=1).font = Font(
            name="微软雅黑",
            bold=True,
            color="17365D",
        )
        ws.cell(row=row_index, column=2).font = Font(name="微软雅黑")
        ws.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True)

    header_row = 12
    headers = ["工作表", "一行代表什么", "数据行数", "用途"]
    for column, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=column, value=value)
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B5")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    purpose = {
        "01_文献": "经 DOI 规范化的文献元数据",
        "02_样品总览": "浏览样品、材料体系、配方摘要和实验变量",
        "03_成分": "按样品查询基体、功能组分、溶剂和含量",
        "04_工艺": "按步骤查询成形、处理和工艺参数",
        "05_结构": "查询 XRD、FTIR、形貌、尺寸、结晶和二级结构",
        "06_性能": "查询力学、热学、电学、传输和物理性能",
        "90_证据与质控": "回查原文证据、样品归属修正、重复合并和复核状态",
    }
    sheet_entries = [
        (
            MASTER_DATA_SHEET,
            "一行代表文献中的一个实际材料样品",
            len(_build_master_rows(dataset)),
            "按原模板顺序直接浏览成分—工艺—结构—性能完整链条",
        ),
        *[
            (
                sheet_name,
                _ROW_MEANINGS[sheet_name],
                len(rows),
                purpose[sheet_name],
            )
            for sheet_name, rows in dataset.sheet_rows().items()
        ],
    ]
    for row_index, values in enumerate(
        sheet_entries,
        start=header_row + 1,
    ):
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EAF2F8")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 58
    ws.freeze_panes = f"A{header_row + 1}"


def _stable_material_data_id(paper_id: str, sample_id: str) -> str:
    """Return a batch-order-independent identifier for one material sample."""

    identity = (
        f"{len(paper_id)}:{paper_id}{len(sample_id)}:{sample_id}"
    ).encode("utf-8")
    return f"MD-{hashlib.sha256(identity).hexdigest()[:24]}"


def _process_route_label(stage: str, method: str) -> tuple[str, str]:
    """Return a readable route label and a stable de-duplication key."""

    normalized_stage = re.sub(r"[\W_]+", "", stage, flags=re.UNICODE).casefold()
    normalized_method = re.sub(
        r"[\W_]+",
        "",
        method,
        flags=re.UNICODE,
    ).casefold()
    if normalized_method == "hme":
        method = "hot-melt extrusion"
        normalized_method = "hotmeltextrusion"
    if method:
        label = (
            method
            if not stage or normalized_stage in _GENERIC_PROCESS_STAGE_KEYS
            else f"{stage}：{method}"
        )
        return label, normalized_method
    return stage, normalized_stage


def _process_parameter_label(
    stage: str,
    parameter: str,
    raw_value: str,
) -> tuple[str, str]:
    """Keep one readable copy of semantically identical process parameters."""

    normalized_stage = re.sub(r"[\W_]+", "", stage, flags=re.UNICODE).casefold()
    display_parameter = re.sub(r"[_\s]+", " ", parameter).strip()
    cleaned_value = _text(_clean_material_chain_summary_value(raw_value))
    prefix = (
        f"[{stage}] "
        if stage and normalized_stage not in _GENERIC_PROCESS_STAGE_KEYS
        else ""
    )
    label = f"{prefix}{display_parameter}={cleaned_value}"
    key = "|".join((
        re.sub(
            r"[\W_]+",
            "",
            display_parameter,
            flags=re.UNICODE,
        ).casefold(),
        re.sub(r"\s+", "", cleaned_value).casefold(),
    ))
    return label, key


def _build_master_rows(dataset: MaterialDataset) -> list[dict[str, Any]]:
    """Project atomic facts into the user's original wide-table reading order.

    The scientific detail sheets remain lossless and atomic.  This first-sheet
    projection is intentionally one row per actual sample so a reader can scan
    the full composition—process—structure—performance chain without seeing
    extraction bookkeeping or generic cross-domain columns first.
    """

    paper_by_id = {row["文献ID"]: row for row in dataset.papers}
    composition_by_sample: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    process_by_sample: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    structure_by_sample: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    performance_by_sample: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    quality_by_sample: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)

    for target, rows in (
        (composition_by_sample, dataset.composition),
        (process_by_sample, dataset.process),
        (structure_by_sample, dataset.structure),
        (performance_by_sample, dataset.performance),
        (quality_by_sample, dataset.quality),
    ):
        for row in rows:
            paper_id = _text(row.get("文献ID"))
            sample_id = _text(row.get("样品ID"))
            if paper_id and sample_id:
                target[(paper_id, sample_id)].append(row)

    master_rows: list[dict[str, Any]] = []
    for sample in dataset.samples:
        paper_id = _text(sample.get("文献ID"))
        sample_id = _text(sample.get("样品ID"))
        key = (paper_id, sample_id)
        paper = paper_by_id.get(paper_id, {})
        composition = composition_by_sample.get(key, [])
        process = sorted(
            process_by_sample.get(key, []),
            key=lambda row: (
                row.get("工序序号") if isinstance(row.get("工序序号"), (int, float)) else 10**9,
                _text(row.get("事实ID")),
            ),
        )
        structure = structure_by_sample.get(key, [])
        performance = performance_by_sample.get(key, [])
        quality = quality_by_sample.get(key, [])
        structure_metric_names = _condition_qualified_metric_names(structure)
        performance_metric_names = _condition_qualified_metric_names(
            performance
        )

        usable_composition = [
            row for row in composition if not _looks_like_misclassified_geometry(row)
        ]
        matrix_rows = [
            row
            for row in usable_composition
            if any(token in _text(row.get("组分角色")) for token in ("基体", "原料", "前驱体"))
        ]
        solvent_rows = [
            row
            for row in usable_composition
            if any(token in _text(row.get("组分角色")) for token in ("溶剂", "助剂"))
        ]
        additive_rows = [
            row
            for row in usable_composition
            if row not in matrix_rows and row not in solvent_rows
        ]

        route_values = []
        parameter_values = []
        post_treatment_values = []
        seen_route_keys: set[str] = set()
        seen_parameter_keys: set[str] = set()
        for row in process:
            stage = _text(row.get("工艺阶段"))
            method = _text(row.get("工艺方法"))
            parameter = _text(row.get("参数名称"))
            raw_value = _measurement_value(row, raw_key="原始值")
            condition = _text(row.get("设备或条件"))
            route_label, route_key = _process_route_label(stage, method)
            if route_label and route_key not in seen_route_keys:
                route_values.append(route_label)
                seen_route_keys.add(route_key)
            if parameter and raw_value:
                parameter_label, parameter_key = _process_parameter_label(
                    stage,
                    parameter,
                    raw_value,
                )
                if parameter_key not in seen_parameter_keys:
                    parameter_values.append(parameter_label)
                    seen_parameter_keys.add(parameter_key)
            if _is_post_treatment_stage(stage):
                post_treatment_values.append(
                    "；".join(part for part in (method, condition) if part)
                )

        # Keep every value self-describing.  Joining names and values in
        # separate de-duplicated lists made repeated measurements ambiguous
        # (for example two contact angles became one name plus two numbers).
        structure_values = [
            "=".join(
                part
                for part in (
                    structure_metric_names[id(row)],
                    _measurement_value(row, raw_key="原始值"),
                )
                if part
            )
            for row in structure
        ]
        performance_values = [
            "=".join(
                part
                for part in (
                    performance_metric_names[id(row)],
                    _measurement_value(row, raw_key="原始值"),
                )
                if part
            )
            for row in performance
        ]
        source_locations = []
        for row in quality:
            location = _clean_source_location(_text(row.get("来源位置")))
            page = _text(row.get("页码"))
            source_locations.append(
                "；".join(
                    part
                    for part in (
                        f"p.{page}" if page else "",
                        location,
                    )
                    if part
                )
            )

        missing_domains = []
        if not usable_composition:
            missing_domains.append("成分")
        if not process:
            missing_domains.append("工艺")
        if not structure:
            missing_domains.append("结构")
        if not performance:
            missing_domains.append("性能")
        if not missing_domains:
            completeness = "完整"
        elif usable_composition and process and (structure or performance):
            completeness = "基本完整"
        else:
            completeness = "不完整"

        result_parts = []
        if structure:
            result_parts.append(
                "结构：" + _join_unique(
                    (
                        f"{structure_metric_names[id(row)]}={_measurement_value(row, raw_key='原始值')}"
                        for row in structure
                    ),
                    limit=1800,
                )
            )
        if performance:
            result_parts.append(
                "性能：" + _join_unique(
                    (
                        f"{performance_metric_names[id(row)]}={_measurement_value(row, raw_key='原始值')}"
                        for row in performance
                    ),
                    limit=1800,
                )
            )

        aliases = _text(sample.get("样品别名"))
        variable = _text(sample.get("主要变量"))
        variable_value = _measurement_parts(
            sample.get("变量值"),
            sample.get("变量单位"),
        )
        note_parts = [
            f"样品别名：{aliases}" if aliases else "",
            f"样品组：{_text(sample.get('样品组'))}" if sample.get("样品组") else "",
            f"主要变量：{variable}={variable_value}" if variable and variable_value else "",
            f"处理状态：{_text(sample.get('处理状态'))}" if sample.get("处理状态") else "",
        ]
        composition_summary_values = [
            (
                f"配方摘要：{_text(sample.get('配方摘要'))}"
                if _text(sample.get("配方摘要"))
                else ""
            ),
            *(
                _component_summary(item, include_role=True)
                for item in usable_composition
            ),
        ]

        row = {
            "数据ID": _stable_material_data_id(paper_id, sample_id),
            "文献编号": paper_id,
            "论文题目": paper.get("文献标题"),
            "DOI|URL": paper.get("DOI"),
            "年份": paper.get("发表年份"),
            "期刊|会议": paper.get("期刊"),
            "材料类别": sample.get("材料体系"),
            "具体材料对象|样品编号": sample_id,
            "纤维形态": sample.get("材料形态"),
            "原料|前驱体|基体": _join_unique(
                (_text(item.get("组分名称")) for item in matrix_rows)
            ) or sample.get("基体"),
            "增强|填料|改性组分": _join_unique(
                (_component_summary(item) for item in additive_rows)
            ),
            "成分配比|浓度": _join_unique(
                composition_summary_values,
                limit=3000,
            ),
            "溶剂|助剂": _join_unique(
                (_component_summary(item) for item in solvent_rows)
            ),
            "工艺路线": _join_unique(route_values, limit=3000),
            "关键工艺参数": _join_unique(parameter_values, limit=6000),
            "后处理条件": _join_unique(post_treatment_values, limit=3000),
            "结构表征方法": _join_unique(
                (_text(item.get("表征方法")) for item in structure)
            ),
            "结构指标名称": _join_unique(
                (structure_metric_names[id(item)] for item in structure)
            ),
            "结构数值": _join_unique(structure_values),
            "结构单位": _join_unique(
                (_text(item.get("单位")) for item in structure)
            ),
            "性能测试方法|标准": _join_unique(
                (_text(item.get("测试方法")) for item in performance)
            ),
            "性能指标名称": _join_unique(
                (performance_metric_names[id(item)] for item in performance)
            ),
            "性能数值": _join_unique(performance_values),
            "性能单位": _join_unique(
                (_text(item.get("单位")) for item in performance)
            ),
            "测试条件": _join_unique(
                (_text(item.get("测试条件")) for item in performance),
                limit=3000,
            ),
            "结果描述|结论": _join_unique(result_parts, limit=4000),
            "数据来源位置": _join_unique(source_locations, limit=3000),
            "原文图表编号": _extract_figure_table_refs(source_locations),
            "是否完整": completeness,
            "缺失信息说明": (
                "当前数据未提取到：" + "、".join(missing_domains)
                if missing_domains
                else ""
            ),
            "备注": _join_unique(note_parts, limit=2000),
        }
        master_rows.append({
            key: _clean_material_chain_summary_value(value)
            for key, value in row.items()
        })

    return master_rows


def _clean_material_chain_summary_value(value: Any) -> Any:
    """Make LLM placeholders readable without changing atomic source facts."""

    if not isinstance(value, str) or not value.strip():
        return value
    text = value.strip()
    text = re.sub(
        r"(?i)\bnot\s+applicable\s*\(([^()]*)\)",
        lambda match: match.group(1).strip(),
        text,
    )
    text = re.sub(
        r"(?i)=\s*not\s+explicitly\s+quantified"
        r"(?:\s*(?:wt\.?\s*%|vol\.?\s*%|at\.?\s*%|%))?",
        "（含量未报告）",
        text,
    )
    text = re.sub(
        r"(?i)\bnot\s+explicitly\s+quantified\b",
        "未报告",
        text,
    )
    text = re.sub(r"(?i)\bnot\s+applicable\b", "", text)
    text = re.sub(
        r"(?i)(?<![\w])"
        r"(?P<unit>wt\.?\s*%|vol\.?\s*%|at\.?\s*%|%|"
        r"°C|K|GPa|MPa|kPa|Pa|rpm|min|h|s|mm|[µμ]m|nm)"
        r"\s+(?P=unit)(?![\w])",
        lambda match: match.group("unit"),
        text,
    )
    text = re.sub(r"；\s*；+", "；", text)
    text = re.sub(r"([：:])\s*；", "；", text)
    return text.strip("； \t")


def build_material_chain_rows(
    dataset: MaterialDataset,
) -> list[dict[str, Any]]:
    """Return the canonical sample-wide rows shared by Excel and platform."""

    return _build_master_rows(dataset)


def _write_master_sheet(
    workbook: Workbook,
    rows: list[dict[str, Any]],
) -> None:
    ws = workbook.create_sheet(MASTER_DATA_SHEET)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "17365D"
    ws.freeze_panes = "G3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 42

    thin = Side(style="thin", color="D5DCE5")
    group_edge = Side(style="medium", color="FFFFFF")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill_by_column: dict[int, str] = {}
    for label, start_column, end_column, dark_color, light_color in _MASTER_GROUPS:
        ws.merge_cells(
            start_row=1,
            start_column=start_column,
            end_row=1,
            end_column=end_column,
        )
        cell = ws.cell(row=1, column=start_column, value=label)
        cell.font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(right=group_edge)
        for column in range(start_column, end_column + 1):
            header_fill_by_column[column] = light_color

    for column, name in enumerate(MASTER_DATA_COLUMNS, start=1):
        cell = ws.cell(row=2, column=column, value=name)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="1F2937")
        cell.fill = PatternFill(
            "solid",
            fgColor=header_fill_by_column.get(column, "E7E6E6"),
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = data_border

    for row_index, row in enumerate(rows, start=3):
        ws.row_dimensions[row_index].height = 66
        for column, name in enumerate(MASTER_DATA_COLUMNS, start=1):
            value = _excel_safe_value(row.get(name))
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.font = Font(name="微软雅黑", size=10, color="1F2937")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = data_border
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7F9FB")

    widths = {
        "数据ID": 20,
        "文献编号": 14,
        "论文题目": 38,
        "DOI|URL": 28,
        "年份": 9,
        "期刊|会议": 28,
        "材料类别": 28,
        "具体材料对象|样品编号": 30,
        "纤维形态": 14,
        "原料|前驱体|基体": 28,
        "增强|填料|改性组分": 30,
        "成分配比|浓度": 38,
        "溶剂|助剂": 24,
        "工艺路线": 38,
        "关键工艺参数": 50,
        "后处理条件": 38,
        "结构表征方法": 28,
        "结构指标名称": 32,
        "结构数值": 26,
        "结构单位": 18,
        "性能测试方法|标准": 30,
        "性能指标名称": 30,
        "性能数值": 24,
        "性能单位": 18,
        "测试条件": 36,
        "结果描述|结论": 48,
        "数据来源位置": 40,
        "原文图表编号": 24,
        "是否完整": 12,
        "缺失信息说明": 30,
        "备注": 38,
    }
    for column, name in enumerate(MASTER_DATA_COLUMNS, start=1):
        letter = ws.cell(row=2, column=column).column_letter
        ws.column_dimensions[letter].width = widths[name]

    end_column = ws.cell(row=2, column=len(MASTER_DATA_COLUMNS)).column_letter
    end_row = max(2, len(rows) + 2)
    ws.auto_filter.ref = f"A2:{end_column}{end_row}"


def _component_summary(
    row: dict[str, Any],
    *,
    include_role: bool = False,
) -> str:
    role = _text(row.get("组分角色"))
    name = _text(row.get("组分名称"))
    amount = _measurement_value(row, raw_key="原始含量")
    prefix = f"{role}：" if include_role and role else ""
    if name and amount:
        return f"{prefix}{name}={amount}"
    return f"{prefix}{name or amount}"


def _measurement_value(
    row: dict[str, Any],
    *,
    raw_key: str,
) -> str:
    raw_value = _text(row.get(raw_key))
    unit = _text(row.get("单位"))
    if raw_value:
        error = row.get("误差")
        if (
            error not in (None, "")
            and re.fullmatch(
                r"[+-]?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?",
                raw_value,
            )
        ):
            return _measurement_parts(
                f"{raw_value} ± {_text(error)}",
                unit,
            )
        return _measurement_parts(raw_value, unit)

    lower = row.get("下限")
    upper = row.get("上限")
    value = row.get("数值")
    error = row.get("误差")
    if lower not in (None, "") or upper not in (None, ""):
        rendered = f"{_text(lower)}–{_text(upper)}".strip("–")
    elif value not in (None, "") and error not in (None, ""):
        rendered = f"{_text(value)} ± {_text(error)}"
    else:
        rendered = _text(value)
    return _measurement_parts(rendered, unit)


def _condition_qualified_metric_names(
    rows: list[dict[str, Any]],
) -> dict[int, str]:
    """Keep repeated measurements paired with their explicit conditions.

    A sample may legitimately report the same metric at several durations or
    test environments.  Flattening names, values and conditions into three
    independent lists made those pairs ambiguous.  Qualify a repeated metric
    only when every distinct value has a non-empty, unique condition; otherwise
    retain the duplicate base name so the strict delivery gate rejects it.
    """

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    rendered = {
        id(row): _text(row.get("指标名称"))
        for row in rows
    }
    for row in rows:
        key = re.sub(
            r"[\W_]+",
            "",
            _text(row.get("指标名称")),
            flags=re.UNICODE,
        ).casefold()
        if key:
            grouped[key].append(row)

    for members in grouped.values():
        distinct_values = {
            _measurement_value(row, raw_key="原始值")
            for row in members
            if _measurement_value(row, raw_key="原始值")
        }
        if len(distinct_values) <= 1:
            continue
        conditions = [
            re.sub(
                r"\s+",
                " ",
                _text(row.get("测试条件")),
            ).strip()
            for row in members
        ]
        normalized_conditions = {
            condition.casefold()
            for condition in conditions
            if condition
        }
        if (
            not all(conditions)
            or len(normalized_conditions) != len(members)
        ):
            continue
        for row, condition in zip(members, conditions):
            base = rendered[id(row)]
            rendered[id(row)] = f"{base} [{condition[:160]}]"
    return rendered


def _measurement_parts(value: Any, unit: Any) -> str:
    rendered_value = _text(value)
    rendered_unit = _text(unit)
    if rendered_value and rendered_unit:
        def normalized(text: str) -> str:
            return (
                re.sub(r"\s+", "", text)
                .casefold()
                .replace("−", "-")
                .replace("⁻", "-")
            )

        if normalized(rendered_value).endswith(normalized(rendered_unit)):
            return rendered_value
        unit_head = re.match(r"^[^\s/·*]+", rendered_unit)
        if unit_head and re.search(
            rf"(?<![\w]){re.escape(unit_head.group(0))}(?![\w])",
            rendered_value,
            flags=re.IGNORECASE,
        ):
            return rendered_value
    return " ".join(part for part in (rendered_value, rendered_unit) if part)


def _join_unique(values: Any, *, limit: int = 8000) -> str:
    seen: set[str] = set()
    rendered: list[str] = []
    for value in values:
        text = _text(value).strip("； ")
        if not text or text in seen:
            continue
        seen.add(text)
        rendered.append(text)
    joined = "；".join(rendered)
    if len(joined) <= limit:
        return joined
    return joined[: max(0, limit - 1)].rstrip("； ") + "…"


def _looks_like_misclassified_geometry(row: dict[str, Any]) -> bool:
    unit = _text(row.get("单位")).lower()
    condition = _text(row.get("条件或说明")).lower()
    return unit in {"nm", "μm", "um", "mm"} and any(
        token in condition for token in ("diameter", "thickness", "直径", "厚度")
    )


def _is_post_treatment_stage(stage: str) -> bool:
    normalized = stage.lower()
    return any(
        token in normalized
        for token in (
            "后处理",
            "交联",
            "干燥",
            "退火",
            "热处理",
            "洗涤",
            "碳化",
            "稳定化",
            "crosslink",
            "dry",
            "anneal",
            "wash",
            "carbon",
        )
    )


def _extract_figure_table_refs(values: Any) -> str:
    references = []
    pattern = re.compile(
        r"(?:\b(?:fig(?:ure)?|table)\.?\s*[A-Za-z]?\d+[A-Za-z]?"
        r"(?:[.\-–—]\w+)?|(?:图|表)\s*[A-Za-z]?\d+[A-Za-z]?)",
        flags=re.IGNORECASE,
    )
    for value in values:
        references.extend(match.group(0) for match in pattern.finditer(_text(value)))
    return _join_unique(references, limit=1000)


def _clean_source_location(value: str) -> str:
    if not value:
        return ""
    cleaned = re.sub(r"\bblock\s+B?\d+\b", "", value, flags=re.IGNORECASE)
    cleaned = re.sub(
        r"\b(?:table_text|paragraph|unknown)\b",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    references = []
    reference_pattern = re.compile(
        r"(?:\bp(?:age)?\.?\s*\d+|\b(?:fig(?:ure)?|table)\.?\s*"
        r"[A-Za-z]?\d+[A-Za-z]?(?:[.\-–—]\w+)?|(?:图|表)\s*"
        r"[A-Za-z]?\d+[A-Za-z]?)",
        flags=re.IGNORECASE,
    )
    references.extend(match.group(0) for match in reference_pattern.finditer(cleaned))
    section_pattern = re.compile(
        r"\b(?:introduction|experimental|methods?|results?|conclusion|"
        r"supplementary(?:\s+information)?)\b",
        flags=re.IGNORECASE,
    )
    references.extend(match.group(0) for match in section_pattern.finditer(cleaned))
    if references:
        return _join_unique(references, limit=1000)
    cleaned = re.sub(r"[\s,;；|]+", " ", cleaned).strip()
    return cleaned[:160]


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _write_sheet(
    workbook: Workbook,
    title: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    color: str,
    table_index: int,
) -> None:
    header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="微软雅黑", size=10, color="1F2937")
    data_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="D9E1F2"),
        right=Side(style="thin", color="D9E1F2"),
        top=Side(style="thin", color="D9E1F2"),
        bottom=Side(style="thin", color="D9E1F2"),
    )

    chunks = [
        rows[index : index + MAX_DATA_ROWS_PER_SHEET]
        for index in range(0, len(rows), MAX_DATA_ROWS_PER_SHEET)
    ] or [[]]
    for chunk_index, chunk in enumerate(chunks, start=1):
        sheet_title = title if chunk_index == 1 else f"{title}_{chunk_index:03d}"
        ws = workbook.create_sheet(sheet_title)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = color
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30
        max_lengths = [len(column) for column in columns]

        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row_idx, row in enumerate(chunk, 2):
            for col_idx, column in enumerate(columns, 1):
                value = _excel_safe_value(row.get(column))
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                if isinstance(value, (int, float)) and column in {
                    "数值",
                    "误差",
                    "下限",
                    "上限",
                    "置信度",
                }:
                    cell.number_format = "0.############"
                if value not in (None, ""):
                    max_lengths[col_idx - 1] = max(
                        max_lengths[col_idx - 1],
                        min(len(str(value)), 60),
                    )

        for col_idx, max_len in enumerate(max_lengths, 1):
            column_name = columns[col_idx - 1]
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            width = max_len + 3
            if column_name in {"证据原文", "测试条件", "设备或条件", "条件或说明"}:
                width = 48
            elif column_name in {"文献标题", "配方摘要", "元数据备注"}:
                width = 42
            elif column_name in {"事实ID", "文献ID"}:
                width = min(max(width, 16), 26)
            else:
                width = min(max(width, 10), 32)
            ws.column_dimensions[column_letter].width = width

        end_column = ws.cell(row=1, column=len(columns)).column_letter
        end_row = max(1, len(chunk) + 1)
        ws.auto_filter.ref = f"A1:{end_column}{end_row}"
        if chunk:
            table_name = f"MaterialTable{table_index:02d}{chunk_index:03d}"
            table = Table(displayName=table_name, ref=f"A1:{end_column}{end_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)


def _excel_safe_value(value: Any) -> Any:
    if isinstance(value, str):
        cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
        if cleaned.lstrip().startswith(("=", "+", "-", "@")):
            return f"'{cleaned}"
        return cleaned
    return value
