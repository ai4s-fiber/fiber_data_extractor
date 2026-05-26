"""Excel export service — generates 数据主表.xlsx with fixed 40 columns."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_excel(rows: list[dict], columns: list[str], filepath: str):
    """Generate a 数据主表.xlsx file with the given rows and column order.

    Args:
        rows: List of dicts, each dict maps column_name -> value.
        columns: Ordered list of 40 column names.
        filepath: Output file path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "数据主表"

    # Header style
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=True)

    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # Auto-adjust column widths (approximate)
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(filepath)
