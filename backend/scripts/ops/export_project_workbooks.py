"""Export resumable per-paper atomic materials workbooks."""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path


BACKEND_ROOT = Path(__file__).resolve().parents[2]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

DEFAULT_REVIEW_STATUSES = [
    "pending",
    "待审核",
    "modified",
    "已修改",
    "approved",
    "通过",
    "uncertain",
    "存疑",
    "missing",
    "缺失",
]
_INVALID_FILENAME_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export complete, resumable per-paper workbooks."
    )
    parser.add_argument("--project-id", type=int, required=True)
    parser.add_argument("--database-url", default="")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument(
        "--review-status",
        action="append",
        default=[],
        help="Candidate review status to include; repeat for multiple values.",
    )
    parser.add_argument(
        "--skip-empty",
        action="store_true",
        help="Do not export completed/review papers that have no selected candidates.",
    )
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--limit", type=int, default=0)
    return parser.parse_args()


def _configure_environment(args: argparse.Namespace) -> None:
    if args.database_url:
        os.environ["DATABASE_URL"] = args.database_url
    os.environ.setdefault("ALLOW_SQLITE_FALLBACK", "true")
    os.environ.setdefault("REDIS_ENABLED", "false")


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _write_json_atomic(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(path)


def _safe_filename_stem(value: str, *, max_length: int = 120) -> str:
    cleaned = _INVALID_FILENAME_RE.sub("_", value).strip(" .")
    cleaned = re.sub(r"\s+", " ", cleaned)
    if not cleaned:
        cleaned = "paper"
    if cleaned.upper() in _WINDOWS_RESERVED_NAMES:
        cleaned = f"_{cleaned}"
    return cleaned[:max_length].rstrip(" .") or "paper"


def _iso(value) -> str:
    if value is None:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc).isoformat()


def _source_signature(
    paper,
    records,
    evidence_items,
    document_blocks,
    fact_candidates,
    sample_catalogs,
) -> str:
    payload = {
        "paper": [paper.id, _iso(paper.updated_at), paper.status],
        "records": [
            [record.id, _iso(record.updated_at), record.review_status]
            for record in records
        ],
        "evidence": [
            [item.id, _iso(item.created_at)]
            for item in evidence_items
        ],
        "blocks": [
            [block.id, _iso(block.created_at)]
            for block in document_blocks
        ],
        "facts": [
            [fact.id, _iso(fact.created_at), fact.assignment_status]
            for fact in fact_candidates
        ],
        "samples": [
            [sample.id, _iso(sample.created_at)]
            for sample in sample_catalogs
        ],
    }
    encoded = json.dumps(
        payload,
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("ascii")
    return hashlib.sha256(encoded).hexdigest()


def _validate_workbook(path: Path, expected_rows: int) -> tuple[bool, str]:
    from openpyxl import load_workbook

    from app.services.workbook_export import WORKBOOK_SHEET_COLUMNS

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return False, f"{exc.__class__.__name__}: {exc}"
    try:
        required = {"00_说明", *WORKBOOK_SHEET_COLUMNS}
        missing = required - set(workbook.sheetnames)
        if missing:
            return False, f"missing_sheets:{','.join(sorted(missing))}"
        for sheet_name, expected_columns in WORKBOOK_SHEET_COLUMNS.items():
            sheet = workbook[sheet_name]
            header = [
                cell.value for cell in next(sheet.iter_rows(max_row=1))
            ]
            if header != expected_columns:
                return False, f"header_mismatch:{sheet_name}"
        if expected_rows:
            core_fact_rows = sum(
                max(0, workbook[sheet_name].max_row - 1)
                for sheet_name in ("03_成分", "04_工艺", "05_结构", "06_性能")
            )
            if core_fact_rows == 0:
                return False, "missing_material_fact_rows"
        return True, ""
    finally:
        workbook.close()


async def _run(args: argparse.Namespace) -> dict:
    import app.models  # noqa: F401
    from sqlalchemy import select

    from app.core.database import async_session_factory, close_database
    from app.models.candidate_record import CandidateRecord
    from app.models.document_parse import DocumentBlock
    from app.models.evidence_item import EvidenceItem
    from app.models.fact_candidate import FactCandidate
    from app.models.paper import Paper
    from app.models.project import Project
    from app.models.sample_catalog import SampleCatalog
    from app.services.workbook_export import generate_structured_workbook

    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "_export_manifest.json"
    previous_entries: dict[str, dict] = {}
    if report_path.is_file():
        try:
            previous_report = json.loads(report_path.read_text(encoding="utf-8"))
            previous_entries = {
                str(item.get("paper_id")): item
                for item in previous_report.get("papers", [])
                if isinstance(item, dict)
            }
        except (OSError, json.JSONDecodeError):
            previous_entries = {}

    statuses = args.review_status or DEFAULT_REVIEW_STATUSES
    report = {
        "project_id": args.project_id,
        "output_dir": str(output_dir),
        "review_statuses": statuses,
        "started_at": _utcnow().isoformat(),
        "updated_at": _utcnow().isoformat(),
        "completed": 0,
        "resumed": 0,
        "skipped_empty": 0,
        "failed": 0,
        "papers": [],
    }
    entries: dict[str, dict] = {}

    try:
        async with async_session_factory() as db:
            project = await db.get(Project, args.project_id)
            if project is None or project.archived_at is not None:
                raise RuntimeError(f"Project {args.project_id} does not exist")
            paper_result = await db.execute(
                select(Paper)
                .where(
                    Paper.project_id == args.project_id,
                    Paper.status.in_(["review", "completed"]),
                )
                .order_by(Paper.id)
            )
            papers = list(paper_result.scalars().all())
            if args.limit > 0:
                papers = papers[: args.limit]

            for index, paper in enumerate(papers, start=1):
                record_result = await db.execute(
                    select(CandidateRecord)
                    .where(
                        CandidateRecord.project_id == args.project_id,
                        CandidateRecord.source_paper_id == paper.id,
                        CandidateRecord.review_status.in_(statuses),
                    )
                    .order_by(CandidateRecord.id)
                )
                records = list(record_result.scalars().all())
                if not records and args.skip_empty:
                    entry = {
                        "paper_id": paper.id,
                        "filename": paper.original_filename,
                        "status": "skipped_empty",
                        "candidate_rows": 0,
                    }
                    entries[str(paper.id)] = entry
                    report["skipped_empty"] += 1
                    continue

                evidence_result = await db.execute(
                    select(EvidenceItem)
                    .where(EvidenceItem.paper_id == paper.id)
                    .order_by(EvidenceItem.id)
                )
                evidence_items = list(evidence_result.scalars().all())
                block_result = await db.execute(
                    select(DocumentBlock)
                    .where(DocumentBlock.paper_id == paper.id)
                    .order_by(
                        DocumentBlock.page_number,
                        DocumentBlock.order_index,
                        DocumentBlock.id,
                    )
                )
                document_blocks = list(block_result.scalars().all())
                fact_result = await db.execute(
                    select(FactCandidate)
                    .where(FactCandidate.paper_id == paper.id)
                    .order_by(FactCandidate.id)
                )
                fact_candidates = list(fact_result.scalars().all())
                sample_result = await db.execute(
                    select(SampleCatalog)
                    .where(SampleCatalog.paper_id == paper.id)
                    .order_by(SampleCatalog.id)
                )
                sample_catalogs = list(sample_result.scalars().all())
                signature = _source_signature(
                    paper,
                    records,
                    evidence_items,
                    document_blocks,
                    fact_candidates,
                    sample_catalogs,
                )
                source_stem = _safe_filename_stem(
                    Path(paper.original_filename).stem
                )
                output_path = (
                    output_dir / f"P{paper.id:06d}_{source_stem}.xlsx"
                )
                previous = previous_entries.get(str(paper.id), {})
                can_resume = (
                    not args.overwrite
                    and output_path.is_file()
                    and previous.get("source_signature") == signature
                )
                if can_resume:
                    valid, validation_error = await asyncio.to_thread(
                        _validate_workbook,
                        output_path,
                        len(records),
                    )
                    if valid:
                        entry = {
                            **previous,
                            "status": "resumed",
                            "validated_at": _utcnow().isoformat(),
                        }
                        entries[str(paper.id)] = entry
                        report["resumed"] += 1
                        print(
                            f"[export] {index}/{len(papers)} resume "
                            f"{output_path.name}",
                            flush=True,
                        )
                        continue

                try:
                    await asyncio.to_thread(
                        generate_structured_workbook,
                        records=records,
                        papers=[paper],
                        evidence_items=evidence_items,
                        document_blocks=document_blocks,
                        fact_candidates=fact_candidates,
                        sample_catalogs=sample_catalogs,
                        filepath=str(output_path),
                    )
                    valid, validation_error = await asyncio.to_thread(
                        _validate_workbook,
                        output_path,
                        len(records),
                    )
                    if not valid:
                        raise RuntimeError(
                            f"Workbook validation failed: {validation_error}"
                        )
                    entry = {
                        "paper_id": paper.id,
                        "filename": paper.original_filename,
                        "workbook_path": str(output_path),
                        "status": "completed",
                        "candidate_rows": len(records),
                        "evidence_rows": len(evidence_items),
                        "fact_rows": len(fact_candidates),
                        "sample_rows": len(sample_catalogs),
                        "parse_blocks": len(document_blocks),
                        "source_signature": signature,
                        "exported_at": _utcnow().isoformat(),
                    }
                    report["completed"] += 1
                    print(
                        f"[export] {index}/{len(papers)} ok "
                        f"{output_path.name} ({len(records)} rows)",
                        flush=True,
                    )
                except Exception as exc:
                    entry = {
                        "paper_id": paper.id,
                        "filename": paper.original_filename,
                        "workbook_path": str(output_path),
                        "status": "failed",
                        "candidate_rows": len(records),
                        "error": f"{exc.__class__.__name__}: {exc}"[:2000],
                    }
                    report["failed"] += 1
                    print(
                        f"[export] {index}/{len(papers)} failed "
                        f"{paper.original_filename}: {exc}",
                        flush=True,
                    )
                entries[str(paper.id)] = entry
                report["papers"] = [
                    entries[key]
                    for key in sorted(entries, key=lambda value: int(value))
                ]
                report["updated_at"] = _utcnow().isoformat()
                await asyncio.to_thread(_write_json_atomic, report_path, report)
    finally:
        await close_database()

    report["papers"] = [
        entries[key] for key in sorted(entries, key=lambda value: int(value))
    ]
    report["healthy"] = report["failed"] == 0
    report["finished_at"] = _utcnow().isoformat()
    report["updated_at"] = report["finished_at"]
    await asyncio.to_thread(_write_json_atomic, report_path, report)
    return report


def main() -> None:
    args = _parse_args()
    _configure_environment(args)
    report = asyncio.run(_run(args))
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if not report["healthy"]:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
