"""Export Review Workbook and strict single-sheet Final 40-field Excel."""
from __future__ import annotations

import json
import sqlite3
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "local_dev_fallback.db"
UPLOAD_DIR = ROOT / "uploads"
DESKTOP = Path(r"C:\Users\Administrator\Desktop")

FINAL_40_HEADERS = [
    "record_id", "paper_id", "paper_title", "doi_or_url", "year", "journal",
    "sample_group_id", "sample_id", "material_system", "fiber_type",
    "variable_name", "variable_value", "variable_unit",
    "composition_expression", "matrix_name", "matrix_content", "matrix_unit",
    "additive_expression", "solvent_or_aid", "composition_evidence",
    "process_route", "spinning_method", "process_parameters", "post_treatment",
    "process_evidence", "structure_methods", "structure_features",
    "structure_evidence", "performance_category", "performance_metric",
    "performance_value", "performance_unit", "performance_method",
    "performance_condition", "performance_evidence", "extraction_method",
    "evidence_text", "ai_confidence", "review_status", "reviewer_comment",
]

DB_FINAL_SELECT = [
    "record_id", "paper_id_biz", "paper_title", "doi_or_url", "year", "journal",
    "sample_group_id", "sample_id", "material_system", "fiber_type",
    "variable_name", "variable_value", "variable_unit",
    "composition_expression", "matrix_name", "matrix_content", "matrix_unit",
    "additive_expression", "solvent_or_aid", "composition_evidence",
    "process_route", "spinning_method", "process_parameters", "post_treatment",
    "process_evidence", "structure_methods", "structure_features",
    "structure_evidence", "performance_category", "performance_metric",
    "performance_value", "performance_unit", "performance_method",
    "performance_condition", "performance_evidence", "extraction_method",
    "evidence_text", "ai_confidence", "review_status", "reviewer_comment",
]

INTERNAL_FORBIDDEN_IN_FINAL = {
    "candidate_status", "source_location", "raw_value", "operator",
    "value_operator", "clean_value", "clean_unit", "range_part",
    "assignment_status", "assignment_confidence", "group_evidence",
}

SAMPLE_MENTION_COLS = [
    "mention_text", "normalized_sample_id", "aliases", "context_text",
    "source_location", "source_type", "confidence",
]
VARIABLE_COLS = [
    "sample_id", "variable_name_raw", "variable_value_raw",
    "variable_unit_raw", "context_text", "source_location", "confidence",
]
GROUP_COLS = [
    "sample_group_id", "sample_ids", "group_variable_name", "group_evidence",
    "source_locations", "confidence", "is_provisional",
]
SAMPLE_CARD_COLS = [
    "sample_id", "sample_aliases", "sample_group_id", "material_system",
    "fiber_type", "variable_name", "variable_value", "variable_unit",
    "composition_expression", "matrix_name", "matrix_content", "matrix_unit",
    "additive_expression", "solvent_or_aid", "composition_evidence",
    "process_route", "spinning_method", "process_parameters", "post_treatment",
    "process_evidence", "structure_methods", "structure_features",
    "structure_evidence", "source_location", "evidence_text", "confidence",
    "_group_confidence", "_group_evidence", "_group_provisional",
]
FACT_COLS = [
    "fact_id", "fact_type", "subject_text", "candidate_sample_ids",
    "metric_or_parameter", "value", "unit", "method", "condition", "category",
    "evidence_text", "source_location", "extraction_method", "confidence",
    "assigned_sample_id", "assignment_confidence", "assignment_status",
]
RESULT_FACT_COLS = [
    "fact_id", "sample_id", "assigned_sample_id", "assignment_status",
    "assignment_confidence", "metric_priority", "raw_metric", "canonical_metric",
    "performance_category", "raw_value", "value_operator", "clean_value",
    "clean_unit", "performance_method", "performance_condition",
    "performance_evidence", "evidence_text", "source_location",
    "extraction_method", "ai_confidence", "export_target", "range_part",
]

MOJIBAKE_REVIEW_STATUS_FIXES = {
    "瀛樼枒": "存疑",
}


def _suffix() -> str:
    if len(sys.argv) <= 1:
        return ""
    value = "".join(ch for ch in sys.argv[1] if ch.isalnum() or ch in ("_", "-"))
    return f"_{value}" if value else ""


def _json_cell(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _clean_review_status(value):
    if not isinstance(value, str):
        return value
    return MOJIBAKE_REVIEW_STATUS_FIXES.get(value, value)


def _load_latest_report() -> dict:
    report_files = sorted(UPLOAD_DIR.glob("*/report_*.json"), key=lambda p: p.stat().st_mtime)
    if not report_files:
        return {}
    with report_files[-1].open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _load_final_rows() -> list[dict]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = []
        cur = conn.execute(f"SELECT {','.join(DB_FINAL_SELECT)} FROM candidate_records ORDER BY id")
        for db_row in cur.fetchall():
            row = dict(db_row)
            row["paper_id"] = row.pop("paper_id_biz", "")
            row["review_status"] = _clean_review_status(row.get("review_status"))
            rows.append({header: row.get(header, "") for header in FINAL_40_HEADERS})
        return rows
    finally:
        conn.close()


def _load_fact_candidates() -> list[dict]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.execute(
            "SELECT fact_id,fact_type,subject_text,candidate_sample_ids,"
            "metric_or_parameter,value,unit,method,condition,category,"
            "evidence_text,source_location,extraction_method,confidence,"
            "assigned_sample_id,assignment_confidence,assignment_status "
            "FROM fact_candidates ORDER BY id"
        )
        return [dict(row) for row in cur.fetchall()]
    finally:
        conn.close()


def assert_final_40_rows(rows: list[dict]) -> None:
    for idx, row in enumerate(rows, 1):
        keys = list(row.keys())
        if keys != FINAL_40_HEADERS:
            raise AssertionError(f"Final row {idx} has invalid columns/order: {keys}")
        extra = INTERNAL_FORBIDDEN_IN_FINAL.intersection(row.keys())
        if extra:
            raise AssertionError(f"Final row {idx} contains internal fields: {sorted(extra)}")


def _write_sheet(wb, title: str, columns: list[str], rows: list[dict], fill: str):
    ws = wb.create_sheet(title)
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(columns, 1):
            value = _json_cell(row.get(col, ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value != "" else None)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(columns, 1):
        max_len = min(max([len(col)] + [
            len(str(row.get(col, ""))) for row in rows[:200] if row.get(col) is not None
        ]), 60)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(10, max_len + 2)
    return ws


def _quality_rows(report: dict, final_rows: list[dict]) -> list[dict]:
    excluded = {
        "sample_mentions", "variable_candidates", "sample_groups",
        "sample_cards", "result_facts", "unassigned_facts",
    }
    rows = [
        {"metric": key, "value": _json_cell(value)}
        for key, value in report.items()
        if key not in excluded
    ]
    rows.extend([
        {"metric": "final_40_columns", "value": ",".join(FINAL_40_HEADERS)},
        {"metric": "final_40_rows", "value": len(final_rows)},
        {"metric": "field_alignment", "value": "pass: strict 40-field export"},
    ])
    return rows


def _write_review_workbook(path: Path, report: dict, final_rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Sample_Mentions", SAMPLE_MENTION_COLS, report.get("sample_mentions", []), "0F766E")
    _write_sheet(wb, "Variable_Candidates", VARIABLE_COLS, report.get("variable_candidates", []), "0369A1")
    _write_sheet(wb, "Sample_Groups", GROUP_COLS, report.get("sample_groups", []), "7C2D12")
    _write_sheet(wb, "Sample_Cards", SAMPLE_CARD_COLS, report.get("sample_cards", []), "059669")
    _write_sheet(wb, "Fact_Candidates", FACT_COLS, _load_fact_candidates(), "A16207")
    _write_sheet(wb, "Result_Facts_QA", RESULT_FACT_COLS, report.get("result_facts", []), "D97706")
    _write_sheet(wb, "Unassigned_Facts", FACT_COLS, report.get("unassigned_facts", []), "BE123C")
    _write_sheet(wb, "Final_Records", FINAL_40_HEADERS, final_rows, "2563EB")
    _write_sheet(wb, "Quality_Report", ["metric", "value"], _quality_rows(report, final_rows), "334155")
    wb.save(path)


def _write_final_40(path: Path, final_rows: list[dict]) -> None:
    assert_final_40_rows(final_rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    _write_sheet_rows(ws, FINAL_40_HEADERS, final_rows)
    wb.save(path)


def _write_sheet_rows(ws, columns: list[str], rows: list[dict]) -> None:
    for col_idx, col in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col)
    for row_idx, row in enumerate(rows, 2):
        keys = list(row.keys())
        if keys != columns:
            raise AssertionError(f"Row {row_idx} has invalid final column order")
        for col_idx, col in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=_json_cell(row.get(col, "")) or None)
    ws.freeze_panes = "A2"


def main() -> None:
    suffix = _suffix()
    review_path = DESKTOP / f"AI4S_Extraction_Review_Workbook{suffix}.xlsx"
    final_path = DESKTOP / f"AI4S_Extraction_Final_40{suffix}.xlsx"
    report = _load_latest_report()
    final_rows = _load_final_rows()
    assert_final_40_rows(final_rows)
    DESKTOP.mkdir(parents=True, exist_ok=True)
    _write_review_workbook(review_path, report, final_rows)
    _write_final_40(final_path, final_rows)
    print(f"Review workbook: {review_path}")
    print(f"Final 40-field Excel: {final_path}")
    print(f"Final rows: {len(final_rows)}")


if __name__ == "__main__":
    main()
