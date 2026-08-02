"""Atomic materials workbook export tests."""

from openpyxl import load_workbook
import pytest

from app.models.candidate_record import CandidateRecord
from app.models.fact_candidate import FactCandidate
from app.models.paper import Paper
from app.models.sample_catalog import SampleCatalog
from app.services.material_data_model import (
    DOMAIN_COMPOSITION,
    DOMAIN_PERFORMANCE,
    DOMAIN_PROCESS,
    DOMAIN_STRUCTURE,
    MaterialDataset,
    _infer_matrix,
    _looks_like_multi_sample_summary,
    _metric_label,
    _normalize_forming_method,
    _sample_specific_additive,
    _sample_specific_process_route,
    build_material_dataset,
    clean_title,
    classify_fact,
)
from app.services.workbook_export import (
    MASTER_DATA_COLUMNS,
    MASTER_DATA_SHEET,
    WORKBOOK_SHEET_COLUMNS,
    build_material_chain_rows,
    _clean_material_chain_summary_value,
    _excel_safe_value,
    _measurement_parts,
    generate_structured_workbook,
)


def test_material_chain_summary_cleans_llm_placeholders():
    assert _clean_material_chain_summary_value(
        "成形：Not applicable (wet papermaking and spraying methods used)"
    ) == "成形：wet papermaking and spraying methods used"
    assert _clean_material_chain_summary_value(
        "基体：PUD=Not explicitly quantified wt.%"
    ) == "基体：PUD（含量未报告）"
    assert _clean_material_chain_summary_value(
        "premixing=homogenized for 10 s s"
    ) == "premixing=homogenized for 10 s"


def test_clean_title_flattens_chemical_formula_markup():
    assert clean_title(
        "Bi<sup>0</sup>-Guided ${ \\mathsf { C o } } "
        "( \\mathsf { O H } ) _ { 2 }$ Catalyst"
    ) == "Bi0-Guided Co(OH)2 Catalyst"


def test_measurement_parts_does_not_duplicate_an_existing_unit():
    assert _measurement_parts("10 to 15% w/v", "% w/v") == "10 to 15% w/v"
    assert _measurement_parts("≈255 %", "%") == "≈255 %"
    assert _measurement_parts("44.89", "%") == "44.89 %"
    assert _measurement_parts(
        "546 mg of Co(NO3)2 precursor",
        "mg of precursor salt",
    ) == "546 mg of Co(NO3)2 precursor"


def test_explicit_performance_type_wins_over_composition_substrings():
    assert _metric_label("Youngs_modulus") == "杨氏模量"

    for metric in (
        "decomposition_rate",
        "soluble_fraction",
        "Poissons_ratio",
    ):
        fact = FactCandidate(
            paper_id=1,
            project_id=1,
            fact_id=f"F-{metric}",
            fact_type="performance",
            metric_or_parameter=metric,
            value="1.0",
            unit="mg mL^-1 h^-1",
            assigned_sample_id="S-1",
            assignment_status="assigned",
            confidence=0.9,
        )
        assert classify_fact(fact) == DOMAIN_PERFORMANCE

    explicit_recipe_fact = FactCandidate(
        paper_id=1,
        project_id=1,
        fact_id="F-silica-content",
        fact_type="performance",
        category="composition",
        metric_or_parameter="silica_content",
        value="15",
        unit="wt%",
        assigned_sample_id="S-1",
        assignment_status="assigned",
        confidence=0.9,
    )
    assert classify_fact(explicit_recipe_fact) == DOMAIN_COMPOSITION

    xps_fact = FactCandidate(
        paper_id=1,
        project_id=1,
        fact_id="F-xps",
        fact_type="performance",
        metric_or_parameter="xps_peak_1",
        value="781",
        unit="eV",
        assigned_sample_id="S-1",
        assignment_status="assigned",
        confidence=0.9,
    )
    assert classify_fact(xps_fact) == DOMAIN_STRUCTURE


def test_component_quantification_does_not_masquerade_as_performance():
    cases = (
        (
            "hap_content",
            "HAp content",
            "20/80 HAp/PCL filament",
            "20",
            "wt%",
        ),
        (
            "pcl_content",
            "PCL content",
            "20/80 HAp/PCL filament",
            "80",
            "wt%",
        ),
        (
            "af_content",
            "AF content",
            "20% AF/PVA aerogel fiber",
            "20",
            "%",
        ),
        (
            "hap_loading_efficiency",
            "HAp loading efficiency",
            "20/80 HAp/PCL filament",
            "87.2 ± 4.1",
            "%",
        ),
        (
            "tga_residue",
            "TGA residue",
            "20/80 HAp/PCL filament",
            "17.4 ± 0.7",
            "wt%",
        ),
    )
    for metric, subject, sample_id, value, unit in cases:
        fact = FactCandidate(
            paper_id=1,
            project_id=1,
            fact_id=f"F-{metric}",
            fact_type="performance",
            category="physical",
            subject_text=subject,
            metric_or_parameter=metric,
            value=value,
            unit=unit,
            evidence_text=(
                "Table 3. TGA results, composition, and HAp loading "
                "efficiency."
            ),
            assigned_sample_id=sample_id,
            assignment_status="assigned",
            confidence=0.9,
        )
        assert classify_fact(fact) == DOMAIN_COMPOSITION


def test_component_quantification_uses_material_roles_in_export():
    sample = SampleCatalog(
        id=1,
        paper_id=1,
        project_id=1,
        sample_id="20/80 HAp/PCL filament",
        material_system="HAp/PCL",
        composition_expression="20 wt% HAp, 80 wt% PCL",
        confidence=0.9,
    )
    facts = [
        FactCandidate(
            paper_id=1,
            project_id=1,
            fact_id="F-hap-content",
            fact_type="performance",
            category="physical",
            metric_or_parameter="hap_content",
            value="20",
            unit="wt%",
            assigned_sample_id=sample.sample_id,
            assignment_status="assigned",
            confidence=0.9,
        ),
        FactCandidate(
            paper_id=1,
            project_id=1,
            fact_id="F-tga-residue",
            fact_type="performance",
            category="physical",
            metric_or_parameter="tga_residue",
            value="17.4",
            unit="wt%",
            assigned_sample_id=sample.sample_id,
            assignment_status="assigned",
            confidence=0.9,
        ),
    ]

    dataset = build_material_dataset(
        records=[],
        papers=[_paper()],
        fact_candidates=facts,
        sample_catalogs=[sample],
        evidence_items=[],
    )

    roles = {
        row["组分名称"]: row["组分角色"]
        for row in dataset.composition
    }
    assert roles["hap_content"] == "配方组分"
    assert roles["tga_residue"] == "实测组成"


def test_die_swell_ratio_is_structure_not_performance():
    for metric, subject in (
        ("d_d_0", "$D/D_0$"),
        ("D/D0", "filament-to-die diameter ratio"),
        ("die-swell ratio", "die swelling ratio"),
    ):
        fact = FactCandidate(
            paper_id=1,
            project_id=1,
            fact_id=f"F-{metric}",
            fact_type="performance",
            category="physical",
            subject_text=subject,
            metric_or_parameter=metric,
            value="1.19",
            unit="",
            evidence_text=(
                "Table 3. Filament average diameter and die swell ratios D/D0."
            ),
            assigned_sample_id="20/80 HAp/PCL filament",
            assignment_status="assigned",
            confidence=0.9,
        )
        assert classify_fact(fact) == DOMAIN_STRUCTURE


def test_polymer_characteristics_do_not_masquerade_as_performance():
    for metric in (
        "m_n_a",
        "mathcal_d_a",
        "surface_roughness",
        "average_grain_size",
        "particle_size",
        "fiber height 1",
        "surface_height",
        "wall_thickness",
        "petal thickness",
        "specific_surface_area",
    ):
        fact = FactCandidate(
            paper_id=1,
            project_id=1,
            fact_id=f"F-{metric}",
            fact_type="performance",
            category="physical",
            metric_or_parameter=metric,
            value="1.0",
            unit="",
            evidence_text=(
                "Table 1. Polymer characteristics and structural results."
            ),
            assigned_sample_id="S-1",
            assignment_status="assigned",
            confidence=0.9,
        )
        assert classify_fact(fact) == DOMAIN_STRUCTURE

    mole_fraction = FactCandidate(
        paper_id=1,
        project_id=1,
        fact_id="F-f-im",
        fact_type="performance",
        category="physical",
        metric_or_parameter="f_text_im_b",
        value="0.29",
        unit="",
        evidence_text=(
            "The mole fraction of NDI-IM (f_IM) was 0.29."
        ),
        assigned_sample_id="S-1",
        assignment_status="assigned",
        confidence=0.9,
    )
    assert classify_fact(mole_fraction) == DOMAIN_COMPOSITION


def test_explicit_process_conditions_and_equipment_dimensions_stay_process():
    for metric, fact_type, method in (
        ("drying_temperature", "process", "oven drying"),
        ("reaction_temperature", "process", "synthesis"),
        (
            "needle_inner_diameter",
            "performance",
            "coaxial spinneret with double concentric needles",
        ),
    ):
        fact = FactCandidate(
            paper_id=1,
            project_id=1,
            fact_id=f"F-{metric}",
            fact_type=fact_type,
            metric_or_parameter=metric,
            value="40",
            unit="°C" if "temperature" in metric else "mm",
            method=method,
            assigned_sample_id="S-1",
            assignment_status="assigned",
            confidence=0.9,
        )
        assert classify_fact(fact) == DOMAIN_PROCESS


def test_tem_requires_a_technique_token_not_temperature_substring():
    actual_tem_peak = FactCandidate(
        paper_id=1,
        project_id=1,
        fact_id="F-tem",
        fact_type="performance",
        metric_or_parameter="particle size",
        method="TEM",
        value="12",
        unit="nm",
        assigned_sample_id="S-1",
        assignment_status="assigned",
        confidence=0.9,
    )
    assert classify_fact(actual_tem_peak) == DOMAIN_STRUCTURE

    temperature = FactCandidate(
        paper_id=1,
        project_id=1,
        fact_id="F-temperature",
        fact_type="process",
        metric_or_parameter="drying_temperature",
        value="40",
        unit="°C",
        assigned_sample_id="S-1",
        assignment_status="assigned",
        confidence=0.9,
    )
    assert classify_fact(temperature) == DOMAIN_PROCESS


def test_matrix_inference_prefers_explicit_polymer_matrix_over_slash_order():
    assert _infer_matrix(
        "AgNFP/PU",
        "AgNFPs (36 vol%) dispersed in a PU matrix",
    ) == "polyurethane (PU)"
    assert _infer_matrix(
        "silk/epoxy resin",
        "unidirectional silk fibers in an epoxy resin matrix",
    ) == "epoxy resin"
    assert _infer_matrix(
        "HAp/PCL",
        "40 wt% HAp and 60 wt% PCL composite filament",
    ) == "polycaprolactone (PCL)"
    assert _infer_matrix(
        "aramid III/PVA",
        "20% aramid III and PVA composite aerogel fiber",
    ) == "polyvinyl alcohol (PVA)"


def test_process_projection_removes_negative_and_cross_sample_summaries():
    assert _normalize_forming_method(
        "Not applicable — this is a wet-chemical precipitation/"
        "self-assembly route, not a fiber spinning process."
    ) == "wet-chemical precipitation/self-assembly route"
    assert _looks_like_multi_sample_summary(
        "Bi precursor was added at varying amounts and concentrations."
    ) is True
    assert _looks_like_multi_sample_summary(
        "Washed and dried at 50 °C. For R-Co(OH)2: electrolysis followed."
    ) is True
    assert _sample_specific_process_route(
        "Solution synthesis followed by drying. "
        "R-Co(OH)2/Bi0.70 prepared via in situ electrolysis.",
        "Co(OH)2",
    ) == "Solution synthesis followed by drying."


def test_targeted_additive_is_kept_only_for_the_named_sample_variant():
    additive = (
        "Silver nanoflake particles (AgNFPs) synthesized from precursors; "
        "EGaIn liquid metal alloy injected into hollow fiber cores"
    )
    assert _sample_specific_additive(
        additive,
        "hollow AgNFP-PU fiber",
        "AgNFPs (36 vol%) in PU matrix, hollow fiber structure",
    ) == "Silver nanoflake particles (AgNFPs) synthesized from precursors"
    assert _sample_specific_additive(
        additive,
        "EGaIn-filled hollow AgNFP-PU fiber",
        "EGaIn alloy injected into the hollow AgNFP-PU fiber",
    ) == additive


def _paper() -> Paper:
    return Paper(
        id=1,
        project_id=1,
        original_filename="fiber.pdf",
        file_object_key="fiber.pdf",
        paper_title="EDC/NHS Crosslinked Electrospun Regenerated Tussah Silk Fibroin Nanofiber Mats",
        doi_or_url="10.1007/s12221-012-0613-y",
        year=2011,
        journal="Fibers and Polymers 2012, Vol.13, No.5",
        status="review",
    )


def _samples() -> list[SampleCatalog]:
    return [
        SampleCatalog(
            id=1,
            paper_id=1,
            project_id=1,
            sample_id="TSF_uncrosslinked",
            sample_aliases='["Uncrosslinked"]',
            sample_group_id="G001",
            material_system="silk fibroin (TSF) nanofibers",
            fiber_type="nanofiber",
            variable_name="treatment",
            variable_value="uncrosslinked",
            composition_expression="electrospun TSF nanofiber mat",
            process_route="electrospinning followed by post-treatment",
            confidence=0.9,
        ),
        SampleCatalog(
            id=2,
            paper_id=1,
            project_id=1,
            sample_id="TSF_ethanol",
            sample_aliases='["Ethanol-treated"]',
            sample_group_id="G001",
            material_system="silk fibroin (TSF) nanofibers",
            fiber_type="nanofiber",
            variable_name="treatment",
            variable_value="ethanol-treated",
            composition_expression="electrospun TSF nanofiber mat treated with ethanol",
            process_route="electrospinning followed by post-treatment",
            confidence=0.9,
        ),
    ]


def _facts() -> list[FactCandidate]:
    return [
        FactCandidate(
            id=1,
            paper_id=1,
            project_id=1,
            fact_id="F-RAW-1",
            fact_type="performance",
            assigned_sample_id="TSF_uncrosslinked",
            metric_or_parameter="fiber_diameter",
            value="787",
            unit="nm",
            method="SEM",
            condition="ethanol-treated",
            evidence_text="After ethanol treatment the diameter increased to 787 nm.",
            confidence=0.92,
            assignment_status="assigned",
        ),
        FactCandidate(
            id=2,
            paper_id=1,
            project_id=1,
            fact_id="F-RAW-2",
            fact_type="performance",
            assigned_sample_id="TSF_ethanol",
            metric_or_parameter="fiber_diameter",
            value="787",
            unit="nm",
            method="SEM",
            condition="ethanol-treated",
            evidence_text="The average diameter was 787 nm.",
            confidence=0.88,
            assignment_status="assigned",
        ),
        FactCandidate(
            id=3,
            paper_id=1,
            project_id=1,
            fact_id="F-RAW-3",
            fact_type="performance",
            assigned_sample_id="TSF_ethanol",
            metric_or_parameter="tensile_strength",
            value="5.51",
            unit="MPa",
            method="tensile test",
            evidence_text="The tensile strength was 5.51 MPa.",
            confidence=0.9,
            assignment_status="assigned",
        ),
    ]


def test_material_dataset_classifies_corrects_and_deduplicates():
    dataset = build_material_dataset(
        records=[],
        papers=[_paper()],
        fact_candidates=_facts(),
        sample_catalogs=_samples(),
        evidence_items=[],
    )

    assert dataset.papers[0]["发表年份"] == 2012
    assert len(dataset.samples) == 2
    assert len(dataset.structure) == 1
    assert dataset.structure[0]["样品ID"] == "TSF_ethanol"
    assert dataset.structure[0]["指标名称"] == "纤维平均直径"
    assert dataset.structure[0]["数值"] == 787.0
    assert len(dataset.performance) == 1
    assert dataset.performance[0]["性能类别"] == "力学性能"
    assert any("重复抽取已合并" in row["质控备注"] for row in dataset.quality)
    assert any("修正样品归属" in row["质控备注"] for row in dataset.quality)


def test_excel_safe_value_removes_control_characters():
    assert _excel_safe_value("20 kV m\u00011") == "20 kV m1"


@pytest.mark.parametrize(
    "value",
    [
        '=HYPERLINK("https://example.invalid")',
        "+SUM(1,2)",
        "-1+2",
        "@SUM(1,2)",
        "  =1+2",
    ],
)
def test_excel_safe_value_neutralizes_formula_injection(value):
    assert _excel_safe_value(value) == f"'{value}"


def test_material_data_id_is_stable_across_sample_order_and_prefix_changes():
    paper_id = "P 0001/中文"
    target_sample_id = "target sample/β"

    def rows_for(sample_ids: list[str]) -> list[dict]:
        dataset = MaterialDataset(
            papers=[{"文献ID": paper_id, "文献标题": "Fiber paper"}],
            samples=[
                {"文献ID": paper_id, "样品ID": sample_id}
                for sample_id in sample_ids
            ],
            composition=[],
            process=[],
            structure=[],
            performance=[],
            quality=[],
        )
        return build_material_chain_rows(dataset)

    def data_id_for(rows: list[dict], sample_id: str) -> str:
        return next(
            row["数据ID"]
            for row in rows
            if row["具体材料对象|样品编号"] == sample_id
        )

    original_id = data_id_for(
        rows_for([target_sample_id, "trailing"]),
        target_sample_id,
    )
    prefixed_id = data_id_for(
        rows_for(["new leading sample", target_sample_id, "trailing"]),
        target_sample_id,
    )
    reordered_id = data_id_for(
        rows_for(["trailing", target_sample_id]),
        target_sample_id,
    )

    assert original_id == prefixed_id == reordered_id
    assert original_id.startswith("MD-")
    digest = original_id.removeprefix("MD-")
    assert len(digest) == 24
    assert set(digest) <= set("0123456789abcdef")
    assert data_id_for(
        rows_for([target_sample_id, "trailing"]),
        "trailing",
    ) != original_id


def test_material_chain_summary_keeps_recipe_context_and_labeled_values():
    dataset = MaterialDataset(
        papers=[{"文献ID": "P-1", "文献标题": "Composite paper"}],
        samples=[
            {
                "文献ID": "P-1",
                "样品ID": "PCL/AA/S",
                "材料体系": "PCL/AA/S BG",
                "配方摘要": (
                    "PCL composite fibers containing undoped 77S glass"
                ),
            }
        ],
        composition=[
            {
                "文献ID": "P-1",
                "样品ID": "PCL/AA/S",
                "组分角色": "基体",
                "组分名称": "PCL",
            }
        ],
        process=[],
        structure=[
            {
                "文献ID": "P-1",
                "样品ID": "PCL/AA/S",
                "指标名称": "fiber diameter",
                "原始值": "500",
                "单位": "nm",
            }
        ],
        performance=[
            {
                "文献ID": "P-1",
                "样品ID": "PCL/AA/S",
                "指标名称": "water contact angle",
                "原始值": "93.4",
                "单位": "°",
            },
            {
                "文献ID": "P-1",
                "样品ID": "PCL/AA/S",
                "指标名称": "water contact angle",
                "原始值": "56.6",
                "单位": "°",
            },
        ],
        quality=[],
    )

    row = build_material_chain_rows(dataset)[0]

    assert row["成分配比|浓度"].startswith(
        "配方摘要：PCL composite fibers containing undoped 77S glass"
    )
    assert "基体：PCL" in row["成分配比|浓度"]
    assert row["结构数值"] == "fiber diameter=500 nm"
    assert row["性能数值"] == (
        "water contact angle=93.4 °；water contact angle=56.6 °"
    )


def test_material_chain_summary_pairs_repeated_metrics_with_conditions():
    dataset = MaterialDataset(
        papers=[{"文献ID": "P-1", "文献标题": "Condition paper"}],
        samples=[{"文献ID": "P-1", "样品ID": "S-1"}],
        composition=[],
        process=[],
        structure=[],
        performance=[
            {
                "文献ID": "P-1",
                "样品ID": "S-1",
                "指标名称": "cell viability",
                "原始值": ">100",
                "单位": "%",
                "测试条件": "1 day extract",
            },
            {
                "文献ID": "P-1",
                "样品ID": "S-1",
                "指标名称": "cell viability",
                "原始值": ">90",
                "单位": "%",
                "测试条件": "3 and 7 day extract",
            },
        ],
        quality=[],
    )

    row = build_material_chain_rows(dataset)[0]

    assert row["性能指标名称"] == (
        "cell viability [1 day extract]；"
        "cell viability [3 and 7 day extract]"
    )
    assert row["性能数值"] == (
        "cell viability [1 day extract]=>100 %；"
        "cell viability [3 and 7 day extract]=>90 %"
    )


def test_material_chain_summary_deduplicates_generic_process_labels():
    dataset = MaterialDataset(
        papers=[{"文献ID": "P-1", "文献标题": "Process paper"}],
        samples=[{"文献ID": "P-1", "样品ID": "S-1"}],
        composition=[],
        process=[
            {
                "文献ID": "P-1",
                "样品ID": "S-1",
                "工艺阶段": "成形",
                "工艺方法": "HME",
                "参数名称": "residence_time",
                "原始值": "10",
                "单位": "s",
                "工序序号": 1,
            },
            {
                "文献ID": "P-1",
                "样品ID": "S-1",
                "工艺阶段": "总体路线",
                "工艺方法": "hot-melt extrusion",
                "参数名称": "residence time",
                "原始值": "10 s",
                "单位": "",
                "工序序号": 2,
            },
            {
                "文献ID": "P-1",
                "样品ID": "S-1",
                "工艺阶段": "后处理",
                "工艺方法": "freeze-drying",
                "参数名称": "duration",
                "原始值": "24",
                "单位": "h",
                "工序序号": 3,
            },
        ],
        structure=[],
        performance=[],
        quality=[],
    )

    row = build_material_chain_rows(dataset)[0]

    assert row["工艺路线"] == "hot-melt extrusion；后处理：freeze-drying"
    assert row["关键工艺参数"] == (
        "residence time=10 s；[后处理] duration=24 h"
    )
    assert row["后处理条件"] == "freeze-drying"


def test_atomic_process_facts_replace_legacy_process_summaries():
    sample = SampleCatalog(
        id=1,
        paper_id=1,
        project_id=1,
        sample_id="20% AF/PVA aerogel fiber",
        material_system="aramid III/PVA",
        fiber_type="aerogel fiber",
        composition_expression="20% AF/PVA",
        process_route="not employed; dried in air",
        confidence=0.88,
    )
    process_fact = FactCandidate(
        paper_id=1,
        project_id=1,
        fact_id="F-process",
        fact_type="process",
        metric_or_parameter="coagulation_bath",
        value="60",
        unit="%",
        method="wet spinning",
        evidence_text="The first coagulation bath contained 60% DMAc/H2O.",
        assigned_sample_id=sample.sample_id,
        assignment_status="assigned",
        confidence=0.9,
    )

    dataset = build_material_dataset(
        records=[],
        papers=[_paper()],
        fact_candidates=[process_fact],
        sample_catalogs=[sample],
        evidence_items=[],
    )
    row = build_material_chain_rows(dataset)[0]

    assert row["工艺路线"] == "wet spinning"
    assert row["关键工艺参数"] == "coagulation bath=60 %"
    assert "not employed" not in row["工艺路线"]


def test_structured_workbook_writes_atomic_material_sheets(tmp_path):
    record = CandidateRecord(
        id=1,
        project_id=1,
        source_paper_id=1,
        record_id="R1",
        paper_id_str="P0001",
        paper_title="Fiber paper",
        sample_id="TSF_ethanol",
        performance_metric="tensile_strength",
        performance_value="5.51",
        performance_unit="MPa",
        evidence_text="The tensile strength was 5.51 MPa.",
        ai_confidence=0.9,
        review_status="待审核",
    )
    output = tmp_path / "result.xlsx"

    generate_structured_workbook(
        records=[record],
        papers=[_paper()],
        evidence_items=[],
        document_blocks=[],
        fact_candidates=_facts(),
        sample_catalogs=_samples(),
        filepath=str(output),
    )

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook.sheetnames == [
            MASTER_DATA_SHEET,
            "00_说明",
            *WORKBOOK_SHEET_COLUMNS.keys(),
        ]
        assert "Main_Data" not in workbook.sheetnames
        master = workbook[MASTER_DATA_SHEET]
        assert master["A1"].value == "文献与样品"
        assert master["J1"].value == "成分"
        assert master["N1"].value == "工艺"
        assert master["Q1"].value == "结构"
        assert master["U1"].value == "性能"
        assert [
            cell.value for cell in next(master.iter_rows(min_row=2, max_row=2))
        ] == MASTER_DATA_COLUMNS
        assert master["B3"].value == "P0001"
        assert master["H3"].value in {"TSF_ethanol", "TSF_uncrosslinked"}
        for sheet_name, columns in WORKBOOK_SHEET_COLUMNS.items():
            sheet = workbook[sheet_name]
            header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
            assert header == columns
        assert workbook["05_结构"]["C2"].value == "TSF_ethanol"
        assert workbook["06_性能"].max_row == 2
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                assert all(
                    not (isinstance(cell.value, str) and cell.value.startswith("="))
                    for cell in row
                )
    finally:
        workbook.close()
    assert not list(tmp_path.glob("*.tmp.xlsx"))
