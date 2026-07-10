"""Regression guardrails using a reference workbook as one fixture (not paper-specific rules)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from app.services.extractor_v7.fact_postprocess import postprocess_extracted_facts
from app.services.extractor_v7.holistic_extract import performances_to_facts
from app.services.extractor_v7.sample_identity import merge_sample_identities
from app.services.extractor_v7.service import V7ExtractorService
from app.services.grouping import build_sample_cards, group_samples

FIXTURE_XLSX = Path(__file__).resolve().parents[2] / "PVDF_recycled_cellulose_fiber_dataset_40fields.xlsx"


@pytest.mark.skipif(not FIXTURE_XLSX.exists(), reason="reference workbook fixture missing")
def test_reference_workbook_baseline_stats():
    wb = openpyxl.load_workbook(FIXTURE_XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    metric_idx = header.index("performance_metric")
    sample_idx = header.index("sample_id")
    metrics = {r[metric_idx] for r in rows if r[metric_idx]}
    samples = {r[sample_idx] for r in rows if r[sample_idx]}
    assert len(rows) >= 50
    assert len(metrics) >= 30
    assert len(samples) >= 10


def test_pipeline_postprocess_improves_exportable_rows_on_coupled_list_fixture():
    """Generic coupled-list expansion should increase exportable performance facts."""
    facts = [
        {
            "fact_id": "F0001",
            "fact_type": "performance",
            "metric_or_parameter": "tensile_strength",
            "value": "",
            "unit": "MPa",
            "evidence_text": "The values of A, B, C composites were about 10, 20, 30 MPa, respectively.",
            "extraction_method": "AI_text",
        }
    ]
    processed, _ = postprocess_extracted_facts(facts, [])
    perf = [f for f in processed if f.get("fact_type") == "performance" and f.get("value")]
    assert len(perf) == 3


def test_sample_identity_merge_increases_card_consolidation():
    mentions = [
        {"normalized_sample_id": "Fiber-1.0wt filler", "aliases": ["F-1.0"]},
        {"normalized_sample_id": "Fiber 1.0wt filler", "aliases": []},
    ]
    facts = [
        {"fact_type": "performance", "assigned_sample_id": "Fiber-1.0wt filler",
         "metric_or_parameter": "tensile_strength", "value": "100"},
    ]
    cards = [
        {"sample_id": "Fiber-1.0wt filler", "sample_aliases": ""},
        {"sample_id": "Fiber 1.0wt filler", "sample_aliases": ""},
    ]
    _, _, merged_cards = merge_sample_identities(mentions, facts, cards)
    assert len(merged_cards) == 1


def test_build_result_facts_exports_numeric_process_parameters_as_qa():
    facts = [
        {
            "fact_id": "F0001",
            "fact_type": "process",
            "metric_or_parameter": "spinning_temperature",
            "value": "150",
            "unit": "°C",
            "assigned_sample_id": "S1",
            "assignment_status": "assigned",
            "assignment_confidence": 0.9,
            "evidence_text": "spinning at 150 °C",
            "source_location": "p.3, Section 2.1",
            "extraction_method": "AI_text",
            "confidence": 0.8,
        }
    ]
    cards = [{"sample_id": "S1", "sample_group_id": "G001", "_group_provisional": False}]
    result_facts = V7ExtractorService._build_result_facts(facts, cards)
    exportable = [rf for rf in result_facts if rf.get("export_target") != "Not exported"]
    assert len(exportable) >= 1
