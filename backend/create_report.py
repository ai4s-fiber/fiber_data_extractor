"""Generate Strong vs Weak comparison Excel report."""
import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
weak_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
strong_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
match_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
miss_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, cols, fill=None):
    f = fill or header_fill
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = header_font
        cell.fill = f
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def write_rows(ws, data, start_row=2, alt_fill=None):
    for r, row_data in enumerate(data, start_row):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if alt_fill and r % 2 == 0:
                cell.fill = alt_fill

# ===== Sheet 1: Overview =====
ws1 = wb.active
ws1.title = 'Overview'
style_header(ws1, 1, ['Metric', 'Weak (mimo-v2.5)', 'Strong (Claude Opus 4.6)'])
overview = [
    ['Candidate records', 33, 58],
    ['Samples identified', 8, 13],
    ['Facts extracted', 49, 58],
    ['Assigned', 33, 58],
    ['Unassigned', 0, 0],
    ['API cost', 'Very low (mimo free)', 'High (proxy paid)'],
    ['Time', '~2 min', '~3 min'],
    ['Model type', 'Weak, needs simple prompt', 'Strong, multi-stage pipeline'],
]
for r, row_data in enumerate(overview, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.border = thin_border
        if c == 2: cell.fill = weak_fill
        elif c == 3: cell.fill = strong_fill
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 34

# ===== Sheet 2: Weak =====
ws2 = wb.create_sheet('Weak-mimo-v2.5')
style_header(ws2, 1, ['#', 'Sample ID', 'Metric', 'Value', 'Unit', 'Category'],
             PatternFill(start_color='548235', end_color='548235', fill_type='solid'))
weak_records = [
    [1, '2MZ-AZINE-PI-30%', 'Imidization degree (ID)', '95.39%', '%', 'thermal'],
    [2, '2MZ-AZINE-PI-30%', 'Imidization degree at 200C', 'essentially complete', '', 'thermal'],
    [3, 'PI aerogel', 'Imidization degree at 300C', 'complete', '', 'thermal'],
    [4, '2MZ-AZINE-PI nanofibers', 'Average fiber diameter', '223.1', 'nm', 'physical'],
    [5, 'PI nanofibers', 'Average fiber diameter', '462.2', 'nm', 'physical'],
    [6, '2MZ-AZINE-PI nanofibers', 'Average fiber length', '40.5', 'um', 'physical'],
    [7, 'PI nanofibers', 'Average fiber length', '22.8', 'um', 'physical'],
    [8, '2MZ-AZINE-PI nanofibers', 'Tensile strength', '2.82', 'MPa', 'mechanical'],
    [9, 'PI nanofibers', 'Tensile strength', '0.72', 'MPa', 'mechanical'],
    [10, 'PI1 aerogel', 'Shrinkage', '30.63%', '%', 'thermal'],
    [11, '2MZ-AZINE-PI3 aerogel', 'Shrinkage', '3.1%', '%', 'thermal'],
    [12, 'PI1 aerogel', 'Density', '12.38', 'mg/cm3', 'physical'],
    [13, '2MZ-AZINE-PI3 aerogel', 'Density', '4.74', 'mg/cm3', 'physical'],
    [14, '2MZ-AZINE-PI3 aerogel', 'Porosity', '99.66%', '%', 'physical'],
    [15, '2MZ-AZINE-PI3 aerogel', 'Water contact angle', '152.7', 'deg', 'physical'],
    [16, 'PI1 aerogel', 'Compressive stress (500 cycles)', '6.14', 'MPa', 'mechanical'],
    [17, 'PI1 aerogel', 'Initial compressive stress', '7.13', 'MPa', 'mechanical'],
    [18, '2MZ-AZINE-PI1 aerogel', 'Thermal conductivity', '26.2', 'mW/m-K', 'thermal'],
    [19, '2MZ-AZINE-PI2 aerogel', 'Thermal conductivity', '25.9', 'mW/m-K', 'thermal'],
    [20, '2MZ-AZINE-PI3 aerogel', 'Thermal conductivity', '25.3', 'mW/m-K', 'thermal'],
    [21, 'PI1 aerogel', 'Thermal conductivity', '26.9', 'mW/m-K', 'thermal'],
    [22, '2MZ-AZINE-PI3 aerogel', 'Upper surface temp (200C)', '62.3', 'C', 'thermal'],
    [23, 'PI1 aerogel', 'Upper surface temp (200C)', '72.8', 'C', 'thermal'],
    [24, '2MZ-AZINE-PI3 aerogel', 'Upper surface temp (400C)', '117.8', 'C', 'thermal'],
    [25, 'PI1 aerogel', 'Upper surface temp (400C)', '150.2', 'C', 'thermal'],
    [26, '2MZ-AZINE-PI3 aerogel', 'Surface temp (80C humid)', '40.5', 'C', 'thermal'],
    [27, 'PI1 aerogel', 'Surface temp (80C humid)', '52.5', 'C', 'thermal'],
    [28, '2MZ-AZINE-PI3 aerogel', 'Real permittivity', '1.004', '', 'dielectric'],
    [29, '2MZ-AZINE-PI3 aerogel', 'Loss tangent', '8e-4', '', 'dielectric'],
    [30, 'PI system', 'Free volume fraction (FFV)', '0.2456', '', 'physical'],
    [31, '2MZ-AZINE-PI system', 'Free volume fraction (FFV)', '0.2535', '', 'physical'],
    [32, 'common PI aerogel', 'Shrinkage', '30.6%', '%', 'thermal'],
    [33, '2MZ-AZINE-PAA solution', 'Electrical conductivity', 'significantly increases', '', 'dielectric'],
]
write_rows(ws2, weak_records, alt_fill=weak_fill)
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 28
ws2.column_dimensions['C'].width = 35
ws2.column_dimensions['D'].width = 22
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 12

# ===== Sheet 3: Strong =====
ws3 = wb.create_sheet('Strong-Claude-Opus')
style_header(ws3, 1, ['#', 'Sample ID', 'Metric', 'Value', 'Unit', 'Category'],
             PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'))
strong_records = [
    [1, '2MZ-AZINE-PI aerogel', 'Density', '4.74', 'mg/cm3', 'physical'],
    [2, '2MZ-AZINE-PI aerogel', 'Water contact angle', '152.7', 'deg', 'physical'],
    [3, 'PI-300C', 'XPS N 1s Imide C-N binding energy', '400.2', 'eV', 'physical'],
    [4, 'PI-200C', 'XPS N 1s Imide C-N binding energy', '400.2', 'eV', 'physical'],
    [5, 'PI-200C', 'XPS N 1s Amide C-N binding energy', '399.6', 'eV', 'physical'],
    [6, 'PI-200C', 'XPS N 1s Imide C-N binding energy', '400.2', 'eV', 'physical'],
    [7, 'PI-200C', 'XPS N 1s Amide C-N binding energy', '399.8', 'eV', 'physical'],
    [8, 'PI-200C', 'XPS N 1s -NH2 binding energy', '399.2', 'eV', 'physical'],
    [9, 'PI-200C', 'XPS N 1s C=N-C binding energy', '398.5', 'eV', 'physical'],
    [10, 'Side Group', 'Time to 98% imidization', '19.6', 'h', 'physical'],
    [11, 'Intra', 'Time to 98% imidization', '20.3', 'h', 'physical'],
    [12, 'Inter', 'Time to 98% imidization', '30.1', 'h', 'physical'],
    [13, '2MZ-AZINE-PI-30%', 'Imidization degree at 200C', '~95-100', '%', 'physical'],
    [14, 'PI-300C', 'Imidization degree at 150C', '~60', '%', 'physical'],
    [15, 'PI-300C', 'Imidization degree at 300C', '~100', '%', 'physical'],
    [16, '2MZ-AZINE-PI-30%', 'Imidization degree at 150C', '~95.39', '%', 'physical'],
    [17, 'PI nanofiber', 'Fiber diameter (mean)', '462.2', 'nm', 'physical'],
    [18, '2MZ-AZINE-PI nanofiber', 'Fiber diameter (mean)', '223.1', 'nm', 'physical'],
    [19, 'PAA', 'Electrical conductivity', '0.443', 'uS/cm', 'dielectric'],
    [20, 'PAA', 'Electrical conductivity', '10.377', 'uS/cm', 'dielectric'],
    [21, 'PI nanofiber', 'Fiber length (mean)', '22.8', 'um', 'physical'],
    [22, '2MZ-AZINE-PI nanofiber', 'Fiber length (mean)', '40.5', 'um', 'physical'],
    [23, '2MZ-AZINE-PI nanofiber', 'Tensile strength at 150C', '2.82', 'MPa', 'physical'],
    [24, 'PI nanofiber', 'Tensile strength at 150C', '0.72', 'MPa', 'physical'],
    [25, 'PI1', 'Shrinkage', '30.63', '%', 'thermal'],
    [26, '2MZ-AZINE-PI1', 'Shrinkage', '0.80', '%', 'thermal'],
    [27, '2MZ-AZINE-PI2', 'Shrinkage', '1.67', '%', 'thermal'],
    [28, '2MZ-AZINE-PI3', 'Shrinkage', '3.13', '%', 'thermal'],
    [29, 'PI1', 'Density', '12.38', 'mg/cm3', 'physical'],
    [30, '2MZ-AZINE-PI1', 'Density', '9.33', 'mg/cm3', 'physical'],
    [31, '2MZ-AZINE-PI2', 'Density', '6.04', 'mg/cm3', 'physical'],
    [32, '2MZ-AZINE-PI3', 'Density', '4.74', 'mg/cm3', 'physical'],
    [33, 'PI1', 'Porosity', '99.13', '%', 'physical'],
    [34, '2MZ-AZINE-PI1', 'Porosity', '99.34', '%', 'physical'],
    [35, '2MZ-AZINE-PI2', 'Porosity', '99.57', '%', 'physical'],
    [36, '2MZ-AZINE-PI3', 'Porosity', '99.66', '%', 'physical'],
    [37, 'PI1', 'Water contact angle', '137.5', 'deg', 'physical'],
    [38, '2MZ-AZINE-PI1', 'Water contact angle', '160.2', 'deg', 'physical'],
    [39, '2MZ-AZINE-PI2', 'Water contact angle', '157.8', 'deg', 'physical'],
    [40, '2MZ-AZINE-PI3', 'Water contact angle', '152.7', 'deg', 'physical'],
    [41, 'PI1', 'Compressive stress (50%, 1st)', '7.13', 'kPa', 'physical'],
    [42, 'PI1', 'Compressive stress (50%, 500th)', '6.14', 'kPa', 'physical'],
    [43, '2MZ-AZINE-PI1', 'Thermal conductivity', '26.2', 'mW/m-K', 'thermal'],
    [44, '2MZ-AZINE-PI2', 'Thermal conductivity', '25.9', 'mW/m-K', 'thermal'],
    [45, '2MZ-AZINE-PI3', 'Thermal conductivity', '25.3', 'mW/m-K', 'thermal'],
    [46, 'PI1', 'Thermal conductivity', '26.9', 'mW/m-K', 'thermal'],
    [47, '2MZ-AZINE-PI3 (200C hot)', 'Upper surface temperature', '62.3', 'C', 'thermal'],
    [48, 'PI1 (200C hot)', 'Upper surface temperature', '72.8', 'C', 'thermal'],
    [49, '2MZ-AZINE-PI3 (400C hot)', 'Upper surface temperature', '117.8', 'C', 'thermal'],
    [50, 'PI1 (400C hot)', 'Upper surface temperature', '150.2', 'C', 'thermal'],
    [51, '2MZ-AZINE-PI3', 'Real permittivity', '1.004', '', 'dielectric'],
    [52, '2MZ-AZINE-PI3', 'Loss tangent', '<8e-4', '', 'dielectric'],
    [53, '2MZ-AZINE-PI system', 'Fractional free volume (FFV)', '0.2456-0.2535', '', 'physical'],
    [54, 'PAA', 'Reaction pathway (Side Group)', '65.0', '%', 'physical'],
    [55, 'Intra', 'Reaction pathway (Intramolecular)', '16.7', '%', 'physical'],
    [56, 'Inter', 'Reaction pathway (Intermolecular)', '18.3', '%', 'physical'],
    [57, 'PI1 (80C humid)', 'Surface temperature after 20min', '52.5', 'C', 'thermal'],
    [58, '2MZ-AZINE-PI3 (80C humid)', 'Surface temperature after 20min', '40.5', 'C', 'thermal'],
]
write_rows(ws3, strong_records, alt_fill=strong_fill)
ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 32
ws3.column_dimensions['C'].width = 38
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 12

# ===== Sheet 4: Cross comparison =====
ws4 = wb.create_sheet('Cross Comparison')
style_header(ws4, 1, ['Metric', 'Weak Sample', 'Weak Value', 'Strong Sample', 'Strong Value', 'Match'],
             PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid'))
cross_data = [
    ['Density (PI1)', 'PI1 aerogel', '12.38 mg/cm3', 'PI1', '12.38 mg/cm3', 'MATCH'],
    ['Density (2MZ-PI3)', '2MZ-AZINE-PI3', '4.74 mg/cm3', '2MZ-AZINE-PI3', '4.74 mg/cm3', 'MATCH'],
    ['Thermal cond. (PI1)', 'PI1 aerogel', '26.9 mW/m-K', 'PI1', '26.9 mW/m-K', 'MATCH'],
    ['Thermal cond. (2MZ-PI3)', '2MZ-AZINE-PI3', '25.3 mW/m-K', '2MZ-AZINE-PI3', '25.3 mW/m-K', 'MATCH'],
    ['Porosity (2MZ-PI3)', '2MZ-AZINE-PI3', '99.66%', '2MZ-AZINE-PI3', '99.66%', 'MATCH'],
    ['Contact angle (2MZ-PI3)', '2MZ-AZINE-PI3', '152.7 deg', '2MZ-AZINE-PI3', '152.7 deg', 'MATCH'],
    ['Shrinkage (PI1)', 'PI1 aerogel', '30.63%', 'PI1', '30.63%', 'MATCH'],
    ['Shrinkage (2MZ-PI3)', '2MZ-AZINE-PI3', '3.1%', '2MZ-AZINE-PI3', '3.13%', 'MATCH'],
    ['Fiber dia. (PI)', 'PI nanofibers', '462.2 nm', 'PI nanofiber', '462.2 nm', 'MATCH'],
    ['Fiber dia. (2MZ-PI)', '2MZ-AZINE-PI', '223.1 nm', '2MZ-AZINE-PI', '223.1 nm', 'MATCH'],
    ['Tensile (2MZ-PI)', '2MZ-AZINE-PI', '2.82 MPa', '2MZ-AZINE-PI', '2.82 MPa', 'MATCH'],
    ['Tensile (PI)', 'PI nanofibers', '0.72 MPa', 'PI nanofiber', '0.72 MPa', 'MATCH'],
    ['Permittivity', '2MZ-AZINE-PI3', '1.004', '2MZ-AZINE-PI3', '1.004', 'MATCH'],
    ['Loss tangent', '2MZ-AZINE-PI3', '8e-4', '2MZ-AZINE-PI3', '<8e-4', 'MATCH'],
    ['Density (2MZ-PI1)', '-', '-', '2MZ-AZINE-PI1', '9.33 mg/cm3', 'Strong only'],
    ['Density (2MZ-PI2)', '-', '-', '2MZ-AZINE-PI2', '6.04 mg/cm3', 'Strong only'],
    ['Porosity (PI1)', '-', '-', 'PI1', '99.13%', 'Strong only'],
    ['Porosity (2MZ-PI1)', '-', '-', '2MZ-AZINE-PI1', '99.34%', 'Strong only'],
    ['Contact angle (PI1)', '-', '-', 'PI1', '137.5 deg', 'Strong only'],
    ['Contact angle (2MZ-PI1)', '-', '-', '2MZ-AZINE-PI1', '160.2 deg', 'Strong only'],
    ['XPS binding energy', '-', '-', 'PI-200/300C', '7 data points', 'Strong only'],
]
for r, row_data in enumerate(cross_data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.border = thin_border
        if 'MATCH' in str(row_data[-1]):
            cell.fill = match_fill
        else:
            cell.fill = miss_fill
ws4.column_dimensions['A'].width = 26
ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 18
ws4.column_dimensions['F'].width = 16

path = r'C:\Users\Administrator\Desktop\AI4S_Strong_vs_Weak_Report.xlsx'
wb.save(path)
print(f'Saved: {path}')
