"""Export candidate_records wide table from DB (weak) + hardcoded (strong) to Excel."""
import sys; sys.stdout.reconfigure(encoding='utf-8')
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=10, color='FFFFFF')
weak_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
strong_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Export columns (match DB schema, skip internal IDs)
EXPORT_COLS = [
    'record_id', 'paper_id_biz', 'paper_title', 'doi_or_url', 'year', 'journal',
    'sample_group_id', 'sample_id', 'material_system', 'fiber_type',
    'variable_name', 'variable_value', 'variable_unit',
    'composition_expression', 'matrix_name', 'matrix_content', 'matrix_unit',
    'additive_expression', 'solvent_or_aid', 'composition_evidence',
    'process_route', 'spinning_method', 'process_parameters', 'post_treatment', 'process_evidence',
    'structure_methods', 'structure_features', 'structure_evidence',
    'performance_category', 'performance_metric', 'performance_value', 'performance_unit',
    'performance_method', 'performance_condition', 'performance_evidence',
    'extraction_method', 'evidence_text', 'ai_confidence',
    'review_status', 'candidate_status', 'source_location',
]


def write_header(ws, fill):
    for c, col_name in enumerate(EXPORT_COLS, 1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border


def write_rows(ws, rows, start_row=2):
    for r, row_data in enumerate(rows, start_row):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val if val != '' else None)
            cell.border = thin_border


def set_widths(ws):
    widths = {
        'A': 12, 'H': 30, 'AE': 40, 'AF': 12,
        'AK': 12, 'AL': 55, 'AN': 10, 'AO': 10, 'AP': 16,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ===== Sheet 1: Weak mode (from DB) =====
ws1 = wb.active
ws1.title = 'Weak-mimo-v2.5'
write_header(ws1, weak_fill)

conn = sqlite3.connect(r'E:\项目\AI4S文献提取\fiber_data_extractor_v6\backend\local_dev_fallback.db')
cur = conn.cursor()
cur.execute(f"SELECT {','.join(EXPORT_COLS)} FROM candidate_records ORDER BY id")
weak_rows = cur.fetchall()
conn.close()

write_rows(ws1, weak_rows)
set_widths(ws1)

# ===== Sheet 2: Strong mode (hardcoded from test output) =====
strong_records = [
    ['V7-S-1', '', '', '', '', '', '', '2MZ-AZINE-PI aerogel', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Density', '4.74', 'mg/cm3', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-2', '', '', '', '', '', '', '2MZ-AZINE-PI aerogel', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Water contact angle', '152.7', 'deg', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-3', '', '', '', '', '', '', 'PI-300C', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'XPS N 1s Imide C-N binding energy', '400.2', 'eV', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-4', '', '', '', '', '', '', 'PI-200C', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'XPS N 1s Imide C-N binding energy', '400.2', 'eV', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-5', '', '', '', '', '', '', 'PI-200C', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'XPS N 1s Amide C-N binding energy', '399.6', 'eV', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-6', '', '', '', '', '', '', 'PI-200C', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'XPS N 1s Imide C-N binding energy', '400.2', 'eV', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-7', '', '', '', '', '', '', 'PI-200C', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'XPS N 1s Amide C-N binding energy', '399.8', 'eV', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-8', '', '', '', '', '', '', 'PI-200C', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'XPS N 1s -NH2 binding energy', '399.2', 'eV', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-9', '', '', '', '', '', '', 'PI-200C', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'XPS N 1s C=N-C binding energy', '398.5', 'eV', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-10', '', '', '', '', '', '', 'Side Group', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Time to 98% imidization', '19.6', 'h', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-11', '', '', '', '', '', '', 'Intra', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Time to 98% imidization', '20.3', 'h', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-12', '', '', '', '', '', '', 'Inter', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Time to 98% imidization', '30.1', 'h', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-13', '', '', '', '', '', '', '2MZ-AZINE-PI-30%', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Imidization degree at 200C', '~95-100', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-14', '', '', '', '', '', '', 'PI-300C', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Imidization degree at 150C', '~60', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-15', '', '', '', '', '', '', 'PI-300C', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Imidization degree at 300C', '~100', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-16', '', '', '', '', '', '', '2MZ-AZINE-PI-30%', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Imidization degree at 150C', '~95.39', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-17', '', '', '', '', '', '', 'PI nanofiber', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Fiber diameter (mean)', '462.2', 'nm', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-18', '', '', '', '', '', '', '2MZ-AZINE-PI nanofiber', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Fiber diameter (mean)', '223.1', 'nm', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-19', '', '', '', '', '', '', 'PAA', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'dielectric', 'Electrical conductivity', '0.443', 'uS/cm', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-20', '', '', '', '', '', '', 'PAA', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'dielectric', 'Electrical conductivity', '10.377', 'uS/cm', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-21', '', '', '', '', '', '', 'PI nanofiber', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Fiber length (mean)', '22.8', 'um', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-22', '', '', '', '', '', '', '2MZ-AZINE-PI nanofiber', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Fiber length (mean)', '40.5', 'um', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-23', '', '', '', '', '', '', '2MZ-AZINE-PI nanofiber', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Tensile strength at 150C', '2.82', 'MPa', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-24', '', '', '', '', '', '', 'PI nanofiber', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Tensile strength at 150C', '0.72', 'MPa', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-25', '', '', '', '', '', '', 'PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Shrinkage', '30.63', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-26', '', '', '', '', '', '', '2MZ-AZINE-PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Shrinkage', '0.80', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-27', '', '', '', '', '', '', '2MZ-AZINE-PI2', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Shrinkage', '1.67', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-28', '', '', '', '', '', '', '2MZ-AZINE-PI3', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Shrinkage', '3.13', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-29', '', '', '', '', '', '', 'PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Density', '12.38', 'mg/cm3', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-30', '', '', '', '', '', '', '2MZ-AZINE-PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Density', '9.33', 'mg/cm3', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-31', '', '', '', '', '', '', '2MZ-AZINE-PI2', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Density', '6.04', 'mg/cm3', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-32', '', '', '', '', '', '', '2MZ-AZINE-PI3', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Density', '4.74', 'mg/cm3', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-33', '', '', '', '', '', '', 'PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Porosity', '99.13', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-34', '', '', '', '', '', '', '2MZ-AZINE-PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Porosity', '99.34', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-35', '', '', '', '', '', '', '2MZ-AZINE-PI2', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Porosity', '99.57', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-36', '', '', '', '', '', '', '2MZ-AZINE-PI3', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Porosity', '99.66', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-37', '', '', '', '', '', '', 'PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Water contact angle', '137.5', 'deg', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-38', '', '', '', '', '', '', '2MZ-AZINE-PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Water contact angle', '160.2', 'deg', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-39', '', '', '', '', '', '', '2MZ-AZINE-PI2', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Water contact angle', '157.8', 'deg', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-40', '', '', '', '', '', '', '2MZ-AZINE-PI3', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Water contact angle', '152.7', 'deg', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-41', '', '', '', '', '', '', 'PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Compressive stress (50%, 1st)', '7.13', 'kPa', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-42', '', '', '', '', '', '', 'PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Compressive stress (50%, 500th)', '6.14', 'kPa', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-43', '', '', '', '', '', '', '2MZ-AZINE-PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Thermal conductivity', '26.2', 'mW/m-K', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-44', '', '', '', '', '', '', '2MZ-AZINE-PI2', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Thermal conductivity', '25.9', 'mW/m-K', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-45', '', '', '', '', '', '', '2MZ-AZINE-PI3', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Thermal conductivity', '25.3', 'mW/m-K', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-46', '', '', '', '', '', '', 'PI1', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Thermal conductivity', '26.9', 'mW/m-K', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-47', '', '', '', '', '', '', '2MZ-AZINE-PI3 (200C hot)', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Upper surface temperature', '62.3', 'C', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-48', '', '', '', '', '', '', 'PI1 (200C hot)', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Upper surface temperature', '72.8', 'C', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-49', '', '', '', '', '', '', '2MZ-AZINE-PI3 (400C hot)', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Upper surface temperature', '117.8', 'C', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-50', '', '', '', '', '', '', 'PI1 (400C hot)', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Upper surface temperature', '150.2', 'C', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-51', '', '', '', '', '', '', '2MZ-AZINE-PI3', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'dielectric', 'Real permittivity', '1.004', '', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-52', '', '', '', '', '', '', '2MZ-AZINE-PI3', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'dielectric', 'Loss tangent', '<8e-4', '', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-53', '', '', '', '', '', '', '2MZ-AZINE-PI system', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Fractional free volume (FFV)', '0.2456-0.2535', '', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-54', '', '', '', '', '', '', 'PAA', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Reaction pathway (Side Group)', '65.0', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-55', '', '', '', '', '', '', 'Intra', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Reaction pathway (Intramolecular)', '16.7', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-56', '', '', '', '', '', '', 'Inter', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'physical', 'Reaction pathway (Intermolecular)', '18.3', '%', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-57', '', '', '', '', '', '', 'PI1 (80C humid)', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Surface temperature after 20min', '52.5', 'C', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
    ['V7-S-58', '', '', '', '', '', '', '2MZ-AZINE-PI3 (80C humid)', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'thermal', 'Surface temperature after 20min', '40.5', 'C', '', '', '', 'AI_text', '', 0.9, 'pending', 'draft', 'results_text'],
]

ws2 = wb.create_sheet('Strong-Claude-Opus')
write_header(ws2, strong_fill)
write_rows(ws2, strong_records)
set_widths(ws2)

path = r'C:\Users\Administrator\Desktop\AI4S_Extracted_Data_Table.xlsx'
wb.save(path)
print(f'Saved: {path}')
print(f'Weak records (from DB): {len(weak_rows)}')
print(f'Strong records: {len(strong_records)}')
print(f'Columns per record: {len(EXPORT_COLS)}')
